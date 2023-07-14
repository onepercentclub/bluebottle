import datetime
import io
import json
import re
from builtins import str
from datetime import timedelta

import dateutil
from django.contrib.gis.geos import Point
from django.test import tag
from django.test.utils import override_settings
from django.urls import reverse
from django.utils.timezone import now
from django_elasticsearch_dsl.test import ESTestCase
from openpyxl import load_workbook
from pytz import UTC
from rest_framework import status

from bluebottle.activities.models import Activity
from bluebottle.activities.serializers import TeamTransitionSerializer
from bluebottle.activities.tests.factories import TeamFactory
from bluebottle.activities.utils import TeamSerializer, InviteSerializer
from bluebottle.collect.tests.factories import CollectActivityFactory, CollectContributorFactory
from bluebottle.deeds.tests.factories import DeedFactory, DeedParticipantFactory
from bluebottle.files.tests.factories import ImageFactory
from bluebottle.funding.tests.factories import FundingFactory, DonorFactory
from bluebottle.initiatives.models import InitiativePlatformSettings, ActivitySearchFilter
from bluebottle.initiatives.tests.factories import InitiativeFactory
from bluebottle.members.models import MemberPlatformSettings
from bluebottle.offices.tests.factories import OfficeSubRegionFactory
from bluebottle.segments.tests.factories import SegmentFactory
from bluebottle.segments.tests.factories import SegmentTypeFactory
from bluebottle.test.factory_models.accounts import BlueBottleUserFactory
from bluebottle.test.factory_models.categories import CategoryFactory
from bluebottle.test.factory_models.geo import LocationFactory, GeolocationFactory, PlaceFactory, CountryFactory
from bluebottle.test.utils import BluebottleTestCase, JSONAPITestClient, APITestCase
from bluebottle.time_based.serializers import PeriodParticipantSerializer
from bluebottle.time_based.tests.factories import (
    DateActivityFactory, PeriodActivityFactory, DateParticipantFactory, PeriodParticipantFactory,
    DateActivitySlotFactory, SkillFactory, TeamSlotFactory
)


@override_settings(
    ELASTICSEARCH_DSL_AUTOSYNC=True,
    ELASTICSEARCH_DSL_AUTO_REFRESH=True
)
@tag('elasticsearch')
class ActivityListSearchAPITestCase(ESTestCase, BluebottleTestCase):
    def setUp(self):
        super(ActivityListSearchAPITestCase, self).setUp()

        self.client = JSONAPITestClient()
        self.url = reverse('activity-preview-list')
        self.owner = BlueBottleUserFactory.create()

    def search(self, filter, sort=None, user=None, place=None, headers=None):
        if isinstance(filter, str):
            url = filter
        else:
            params = dict(
                (f'filter[{key}]', value) for key, value in filter.items()
            )

            if sort:
                params['sort'] = sort

            if place:
                params['place'] = place

            query = '&'.join(f'{key}={value}' for key, value in params.items())

            url = f'{self.url}?{query}'

        if headers is None:
            headers = {}

        response = self.client.get(
            url,
            user=user,
            **headers
        )

        self.data = json.loads(response.content)

    def assertFound(self, matching, count=None):
        self.assertEqual(self.data['meta']['pagination']['count'], len(matching))

        if count:
            self.assertEqual(len(self.data['data']), count)

        ids = set(str(activity.pk) for activity in matching)

        for activity in self.data['data']:
            self.assertTrue(activity['id'] in ids)

    def assertFacets(self, filter, facets):
        counts = dict(
            (facet['id'], facet['count']) for facet in self.data['meta']['facets'][filter]
        )
        names = dict(
            (facet['id'], facet.get('name')) for facet in self.data['meta']['facets'][filter]
        )

        for key, (name, value) in facets.items():
            self.assertEqual(counts[key], value)
            self.assertEqual(names[key], name)

    def test_images(self):
        DateActivityFactory.create(
            owner=self.owner, status='open', image=ImageFactory.create()
        )
        PeriodActivityFactory.create(status='open', image=ImageFactory.create())
        FundingFactory.create(review_status='open', image=ImageFactory.create())

        response = self.client.get(self.url, user=self.owner)

        for activity in response.json()['data']:
            self.assertTrue(
                re.match('^/api/activities/\d+/image/600x337', activity['attributes']['image'])
            )

    def test_deed_preview(self):
        activity = DeedFactory.create(status='open')
        response = self.client.get(self.url, user=self.owner)
        attributes = response.json()['data'][0]['attributes']

        self.assertEqual(attributes['slug'], activity.slug)
        self.assertEqual(attributes['title'], activity.title)
        self.assertEqual(attributes['initiative'], activity.initiative.title)
        self.assertEqual(attributes['status'], activity.status)
        self.assertEqual(attributes['team-activity'], activity.team_activity)
        self.assertEqual(attributes['is-online'], None)
        self.assertEqual(attributes['is-full'], None)
        self.assertEqual(attributes['theme'], activity.initiative.theme.name)

    def test_date_preview(self):
        activity = DateActivityFactory.create(status='open')
        response = self.client.get(self.url, user=self.owner)
        attributes = response.json()['data'][0]['attributes']

        self.assertEqual(attributes['slug'], activity.slug)
        self.assertEqual(attributes['title'], activity.title)
        self.assertEqual(attributes['initiative'], activity.initiative.title)
        self.assertEqual(attributes['status'], activity.status)
        self.assertEqual(attributes['team-activity'], activity.team_activity)
        self.assertEqual(attributes['is-online'], False)
        self.assertEqual(attributes['is-full'], False)
        self.assertEqual(attributes['theme'], activity.initiative.theme.name)
        self.assertEqual(attributes['expertise'], activity.expertise.name)
        self.assertEqual(attributes['slot-count'], 1)
        self.assertEqual(dateutil.parser.parse(attributes['start']), activity.slots.first().start)
        self.assertEqual(dateutil.parser.parse(attributes['end']), activity.slots.first().end)
        self.assertEqual(attributes['has-multiple-locations'], False)
        location = activity.slots.first().location
        self.assertEqual(
            attributes['location'], f'{location.locality}, {location.country.alpha2_code}'
        )

    def test_date_preview_multiple_slots(self):
        activity = DateActivityFactory.create(status='open', slots=[])
        DateActivitySlotFactory.create_batch(3, activity=activity)
        DateActivitySlotFactory.create(
            status='succeeded', activity=activity, start=now() - timedelta(days=10)
        )
        response = self.client.get(self.url + '?filter[upcoming]=1', user=self.owner)
        attributes = response.json()['data'][0]['attributes']

        self.assertEqual(attributes['slug'], activity.slug)
        self.assertEqual(attributes['title'], activity.title)
        self.assertEqual(attributes['initiative'], activity.initiative.title)
        self.assertEqual(attributes['status'], activity.status)
        self.assertEqual(attributes['team-activity'], activity.team_activity)
        self.assertEqual(attributes['is-online'], False)
        self.assertEqual(attributes['is-full'], False)
        self.assertEqual(attributes['theme'], activity.initiative.theme.name)
        self.assertEqual(attributes['expertise'], activity.expertise.name)
        self.assertEqual(attributes['slot-count'], 3)
        self.assertEqual(
            dateutil.parser.parse(attributes['start']),
            min([slot.start for slot in activity.slots.all() if slot.start > now()])
        )

        self.assertEqual(
            dateutil.parser.parse(attributes['end']),
            min([slot.end for slot in activity.slots.all() if slot.start > now()])
        )
        self.assertEqual(attributes['has-multiple-locations'], True)
        self.assertIsNone(attributes['location'])

    def test_date_preview_multiple_slots_single_location(self):
        activity = DateActivityFactory.create(status='open', slots=[])
        location = GeolocationFactory.create()
        DateActivitySlotFactory.create_batch(3, activity=activity, location=location)
        response = self.client.get(self.url, user=self.owner)
        attributes = response.json()['data'][0]['attributes']

        self.assertEqual(attributes['slug'], activity.slug)
        self.assertEqual(attributes['title'], activity.title)
        self.assertEqual(attributes['initiative'], activity.initiative.title)
        self.assertEqual(attributes['status'], activity.status)
        self.assertEqual(attributes['team-activity'], activity.team_activity)
        self.assertEqual(attributes['is-online'], False)
        self.assertEqual(attributes['is-full'], False)
        self.assertEqual(attributes['theme'], activity.initiative.theme.name)
        self.assertEqual(attributes['expertise'], activity.expertise.name)
        self.assertEqual(attributes['slot-count'], 3)
        self.assertEqual(attributes['has-multiple-locations'], False)

        self.assertEqual(
            attributes['location'], f'{location.locality}, {location.country.alpha2_code}'
        )

    def test_date_preview_multiple_slots_single_open(self):
        activity = DateActivityFactory.create(status='open', slots=[])
        DateActivitySlotFactory.create(activity=activity, status='draft', is_online=None)
        open_slot = DateActivitySlotFactory.create(activity=activity)

        response = self.client.get(self.url, user=self.owner)
        attributes = response.json()['data'][0]['attributes']

        self.assertEqual(attributes['slot-count'], 1)
        self.assertEqual(attributes['has-multiple-locations'], False)

        self.assertEqual(
            attributes['location'],
            f'{open_slot.location.locality}, {open_slot.location.country.alpha2_code}'
        )

        self.assertEqual(dateutil.parser.parse(attributes['start']), open_slot.start)
        self.assertEqual(dateutil.parser.parse(attributes['end']), open_slot.end)

    def test_date_preview_multiple_slots_filtered(self):
        activity = DateActivityFactory.create(status='open', slots=[])
        DateActivitySlotFactory.create(activity=activity, start=now() + timedelta(days=14))
        DateActivitySlotFactory.create(activity=activity, start=now() + timedelta(days=21))
        current_slot = DateActivitySlotFactory.create(activity=activity, start=now() + timedelta(days=7))

        start = now()
        end = start + timedelta(days=12)
        response = self.client.get(
            self.url + '?filter[date]={},{}'.format(
                start.strftime('%Y-%m-%d'),
                end.strftime('%Y-%m-%d')
            )
        )
        attributes = response.json()['data'][0]['attributes']
        self.assertEqual(attributes['slot-count'], 1)

        self.assertEqual(attributes['has-multiple-locations'], False)
        self.assertEqual(attributes['is-online'], False)
        self.assertEqual(dateutil.parser.parse(attributes['start']), current_slot.start)
        self.assertEqual(dateutil.parser.parse(attributes['end']), current_slot.end)

        location = current_slot.location
        self.assertEqual(
            attributes['location'], f'{location.locality}, {location.country.alpha2_code}'
        )

    def test_date_preview_multiple_slots_succeeded(self):
        activity = DateActivityFactory.create(slots=[])
        first = DateActivitySlotFactory.create(activity=activity, start=now() - timedelta(days=21))
        DateActivitySlotFactory.create(activity=activity, start=now() - timedelta(days=14))
        last = DateActivitySlotFactory.create(activity=activity, start=now() - timedelta(days=7))

        activity.status = 'succeeded'
        activity.save()

        response = self.client.get(self.url)
        attributes = response.json()['data'][0]['attributes']
        self.assertEqual(attributes['slot-count'], 0)

        self.assertEqual(attributes['has-multiple-locations'], True)
        self.assertEqual(attributes['is-online'], False)

        self.assertEqual(dateutil.parser.parse(attributes['start']), first.start)
        self.assertEqual(dateutil.parser.parse(attributes['end']), last.end)

    def test_date_preview_all_full(self):
        activity = DateActivityFactory.create(status='open', slots=[])
        DateActivitySlotFactory.create_batch(3, activity=activity, status='full')
        response = self.client.get(self.url, user=self.owner)
        attributes = response.json()['data'][0]['attributes']
        self.assertEqual(attributes['is-full'], True)

    def test_date_preview_is_online(self):
        activity = DateActivityFactory.create(status='open', slots=[])
        DateActivitySlotFactory.create_batch(
            3, activity=activity, location=None, is_online=True, status='full'
        )
        response = self.client.get(self.url, user=self.owner)
        attributes = response.json()['data'][0]['attributes']
        self.assertEqual(attributes['is-online'], True)

    def test_date_preview_matching(self):
        activity = DateActivityFactory.create(
            status='open',
            slots=[]
        )
        DateActivitySlotFactory.create(
            activity=activity,
            location=GeolocationFactory.create(position=Point(20.1, 10.1))
        )

        DateActivitySlotFactory.create(
            activity=activity
        )

        self.owner.favourite_themes.add(activity.initiative.theme)
        self.owner.skills.add(activity.expertise)
        self.owner.place = PlaceFactory.create(
            position=Point(20.0, 10.0)
        )
        self.owner.save()

        response = self.client.get(self.url, user=self.owner)
        attributes = response.json()['data'][0]['attributes']
        self.assertEqual(attributes['matching-properties']['theme'], True)
        self.assertEqual(attributes['matching-properties']['skill'], True)
        self.assertEqual(attributes['matching-properties']['location'], True)

    def test_period_preview(self):
        activity = PeriodActivityFactory.create(status='open', is_online=False)
        response = self.client.get(self.url, user=self.owner)
        attributes = response.json()['data'][0]['attributes']

        self.assertEqual(attributes['slug'], activity.slug)
        self.assertEqual(attributes['title'], activity.title)
        self.assertEqual(attributes['initiative'], activity.initiative.title)
        self.assertEqual(attributes['status'], activity.status)
        self.assertEqual(attributes['team-activity'], activity.team_activity)
        self.assertEqual(attributes['is-online'], False)
        self.assertEqual(attributes['is-full'], None)
        self.assertEqual(attributes['theme'], activity.initiative.theme.name)
        self.assertEqual(attributes['expertise'], activity.expertise.name)
        self.assertEqual(attributes['slot-count'], None)
        self.assertEqual(attributes['has-multiple-locations'], False)
        self.assertEqual(attributes['contribution-duration'], {'period': 'overall', 'value': 20.0})

        location = activity.location
        self.assertEqual(
            attributes['location'], f'{location.locality}, {location.country.alpha2_code}'
        )

        self.assertEqual(attributes['matching-properties']['theme'], False)
        self.assertEqual(attributes['matching-properties']['skill'], False)
        self.assertEqual(attributes['matching-properties']['location'], False)

    def test_period_preview_matching(self):
        activity = PeriodActivityFactory.create(
            status='open',
            location=GeolocationFactory.create(position=Point(20.1, 10.1))
        )

        self.owner.favourite_themes.add(activity.initiative.theme)
        self.owner.skills.add(activity.expertise)
        self.owner.place = PlaceFactory.create(
            position=Point(20.0, 10.0)
        )
        self.owner.save()

        response = self.client.get(self.url, user=self.owner)
        attributes = response.json()['data'][0]['attributes']
        self.assertEqual(attributes['matching-properties']['theme'], True)
        self.assertEqual(attributes['matching-properties']['skill'], True)
        self.assertEqual(attributes['matching-properties']['location'], True)

    def test_funding_preview(self):
        activity = FundingFactory.create(status='open')
        response = self.client.get(self.url, user=self.owner)
        attributes = response.json()['data'][0]['attributes']

        self.assertEqual(attributes['slug'], activity.slug)
        self.assertEqual(attributes['title'], activity.title)
        self.assertEqual(attributes['initiative'], activity.initiative.title)
        self.assertEqual(attributes['status'], activity.status)
        self.assertEqual(attributes['team-activity'], activity.team_activity)
        self.assertEqual(attributes['is-online'], None)
        self.assertEqual(attributes['is-full'], None)
        self.assertEqual(attributes['theme'], activity.initiative.theme.name)
        self.assertEqual(attributes['expertise'], None)
        self.assertEqual(attributes['slot-count'], None)
        self.assertEqual(attributes['has-multiple-locations'], False)

        location = activity.initiative.place
        self.assertEqual(
            attributes['location'], f'{location.locality}, {location.country.alpha2_code}'
        )

    def test_collect_preview(self):
        activity = CollectActivityFactory.create(status='open')
        response = self.client.get(self.url, user=self.owner)
        attributes = response.json()['data'][0]['attributes']

        self.assertEqual(attributes['slug'], activity.slug)
        self.assertEqual(attributes['title'], activity.title)
        self.assertEqual(attributes['initiative'], activity.initiative.title)
        self.assertEqual(attributes['status'], activity.status)
        self.assertEqual(attributes['team-activity'], activity.team_activity)
        self.assertEqual(attributes['is-online'], None)
        self.assertEqual(attributes['is-full'], None)
        self.assertEqual(attributes['theme'], activity.initiative.theme.name)
        self.assertEqual(attributes['expertise'], None)
        self.assertEqual(attributes['slot-count'], None)
        self.assertEqual(attributes['has-multiple-locations'], False)
        self.assertEqual(attributes['collect-type'], activity.collect_type.name)

        location = activity.location
        self.assertEqual(
            attributes['location'], f'{location.locality}, {location.country.alpha2_code}'
        )

    def test_collect_preview_dutch(self):
        activity = CollectActivityFactory.create(status='open')
        theme_translation = activity.initiative.theme.translations.get(
            language_code='nl'
        )

        collect_type_translation = activity.collect_type.translations.get(
            language_code='nl'
        )
        response = self.client.get(self.url, HTTP_X_APPLICATION_LANGUAGE='nl')
        attributes = response.json()['data'][0]['attributes']

        self.assertEqual(attributes['theme'], theme_translation.name)
        self.assertEqual(attributes['collect-type'], collect_type_translation.name)

    def test_search(self):
        text = 'consectetur adipiscing elit,'
        title = PeriodActivityFactory.create(
            status="open", title=f'title with {text}',
        )
        description = PeriodActivityFactory.create(
            status="open", description=f'description with {text}',
        )

        initiative_title = PeriodActivityFactory.create(
            status="open", initiative=InitiativeFactory.create(title=f'title with {text}'),
        )
        initiative_story = PeriodActivityFactory.create(
            status="open", initiative=InitiativeFactory.create(story=f'story with {text}'),
        )

        initiative_pitch = PeriodActivityFactory.create(
            status="open", initiative=InitiativeFactory.create(pitch=f'pitch with {text}'),
        )

        slot_title = DateActivityFactory.create(status="open")
        DateActivitySlotFactory.create(activity=slot_title, title=f'slot title with {text}')

        response = self.client.get(
            f'{self.url}?filter[search]={text}',
        )

        data = json.loads(response.content)
        ids = [int(activity['id']) for activity in data['data']]

        self.assertTrue(title.pk in ids)
        self.assertTrue(description.pk in ids)
        self.assertTrue(initiative_title.pk in ids)
        self.assertTrue(initiative_pitch.pk in ids)
        self.assertTrue(initiative_story.pk in ids)
        self.assertTrue(slot_title.pk in ids)

        self.assertEqual(data['meta']['pagination']['count'], 6)

    def test_sort_upcoming(self):
        today = now().date()
        first_date_activity = DateActivityFactory.create(status='open', slots=[])
        second_date_activity = DateActivityFactory.create(status='open', slots=[])
        activities = [
            PeriodActivityFactory(
                status='open', start=None, deadline=now() + timedelta(days=1)
            ),
            PeriodActivityFactory(
                status='open', start=None, deadline=None
            ),

            PeriodActivityFactory(
                status='open', start=now() - timedelta(days=1), deadline=None
            ),

            first_date_activity,
            second_date_activity,

            PeriodActivityFactory(status='open', start=today + timedelta(days=8)),
            CollectActivityFactory(status='open', start=today + timedelta(days=9)),
        ]

        DateActivitySlotFactory.create(
            status='open', start=now() + timedelta(days=2), activity=first_date_activity
        )
        DateActivitySlotFactory.create(
            status='open', start=now() + timedelta(days=5), activity=first_date_activity
        )
        DateActivitySlotFactory.create(
            status='open', start=now() - timedelta(days=5), activity=first_date_activity
        )

        DateActivitySlotFactory.create(
            status='open', start=now() + timedelta(days=4), activity=second_date_activity
        )
        DateActivitySlotFactory.create(
            status='open', start=now() + timedelta(days=7), activity=second_date_activity
        )
        DateActivitySlotFactory.create(
            status='open', start=now() - timedelta(days=7), activity=second_date_activity
        )

        self.search({'upcoming': 1})

        self.assertEqual(
            [str(activity.pk) for activity in activities],
            [activity['id'] for activity in self.data['data']]
        )

    def test_sort_unknown(self):
        activities = [
            PeriodActivityFactory(
                status='open', start=None, deadline=now() + timedelta(days=1)
            ),
            PeriodActivityFactory(
                status='open', start=None, deadline=None
            ),

            PeriodActivityFactory(
                status='open', start=now() - timedelta(days=1), deadline=None
            ),

        ]

        self.search({}, sort='some-unknown-sort-option')

        self.assertEqual(
            [str(activity.pk) for activity in activities],
            [activity['id'] for activity in self.data['data']]
        )

    def test_sort_upcoming_false(self):
        today = now().date()
        first_date_activity = DateActivityFactory.create(status='succeeded', slots=[])
        second_date_activity = DateActivityFactory.create(status='succeeded', slots=[])
        activities = [

            CollectActivityFactory(
                status='succeeded',
                start=today - timedelta(days=10),
                end=today - timedelta(days=1)
            ),

            second_date_activity,
            first_date_activity,

            PeriodActivityFactory(
                status='succeeded', start=None, deadline=now() - timedelta(days=6)
            ),

            PeriodActivityFactory(
                status='succeeded', start=None, deadline=now() - timedelta(days=10)
            ),

        ]

        DateActivitySlotFactory.create(
            status='finished', start=now() - timedelta(days=4), activity=first_date_activity
        )

        DateActivitySlotFactory.create(
            status='finished', start=now() - timedelta(days=5), activity=second_date_activity
        )
        DateActivitySlotFactory.create(
            status='finished', start=now() - timedelta(days=2), activity=second_date_activity
        )

        self.search({'upcoming': 0})

        self.assertEqual(
            [str(activity.pk) for activity in activities],
            [activity['id'] for activity in self.data['data']]
        )

    def test_sort_distance(self):
        amsterdam = GeolocationFactory.create(position=Point(4.922114, 52.362438))
        leiden = GeolocationFactory.create(position=Point(4.491056, 52.166758))
        texel = GeolocationFactory.create(position=Point(4.853281, 53.154617))
        lyutidol = GeolocationFactory.create(position=Point(23.676222, 43.068555))

        activity_amsterdam = PeriodActivityFactory(status="open", location=amsterdam)
        activity_online1 = PeriodActivityFactory(status="open", is_online=True)
        activity_leiden = PeriodActivityFactory(status="open", location=leiden)
        activity_texel = PeriodActivityFactory(status="open", location=texel)
        activity_online2 = PeriodActivityFactory(status="open", is_online=True)
        activity_lyutidol = PeriodActivityFactory(status="open", location=lyutidol)

        leiden_place = PlaceFactory.create(position=leiden.position)
        texel_place = PlaceFactory.create(position=texel.position)

        self.search(
            filter={'distance': '500km', 'is_online': '0'},
            sort='distance',
            place=leiden_place.pk
        )
        data = self.data['data']
        self.assertEqual(data[0]['id'], str(activity_leiden.id))
        self.assertEqual(data[1]['id'], str(activity_amsterdam.id))
        self.assertEqual(data[2]['id'], str(activity_texel.id))
        self.assertEqual(len(data), 3)

        # Widen search and search from Texel
        self.search(
            filter={'distance': '5000km', 'is_online': '0'},
            sort='distance',
            place=texel_place.pk
        )
        data = self.data['data']
        self.assertEqual(data[0]['id'], str(activity_texel.id))
        self.assertEqual(data[1]['id'], str(activity_amsterdam.id))
        self.assertEqual(data[2]['id'], str(activity_leiden.id))
        self.assertEqual(data[3]['id'], str(activity_lyutidol.id))
        self.assertEqual(len(data), 4)

        # With online
        self.search(
            filter={'distance': '500km'},
            sort='distance',
            place=leiden_place.pk
        )
        data = self.data['data']
        self.assertEqual(len(data), 5)
        self.assertEqual(data[0]['id'], str(activity_online1.id))
        self.assertEqual(data[1]['id'], str(activity_online2.id))
        self.assertEqual(data[2]['id'], str(activity_leiden.id))
        self.assertEqual(data[3]['id'], str(activity_amsterdam.id))
        self.assertEqual(data[4]['id'], str(activity_texel.id))

        # Any distance
        self.search(
            filter={'is_online': '0'},
            sort='distance',
            place=leiden_place.pk
        )
        data = self.data['data']
        self.assertEqual(data[0]['id'], str(activity_leiden.id))
        self.assertEqual(data[1]['id'], str(activity_amsterdam.id))
        self.assertEqual(data[2]['id'], str(activity_texel.id))
        self.assertEqual(data[3]['id'], str(activity_lyutidol.id))
        self.assertEqual(len(data), 4)

    def test_sort_date(self):
        today = now().date()
        activities = [
            PeriodActivityFactory(status='open', start=now() - timedelta(days=1), deadline=now() + timedelta(days=10)),
            DateActivityFactory.create(status='open', slots=[]),
            DateActivityFactory.create(status='open', slots=[]),
            PeriodActivityFactory(status='open', start=today + timedelta(days=8)),
            CollectActivityFactory(status='open', start=today + timedelta(days=9)),
        ]
        DateActivitySlotFactory.create(status='open', start=now() + timedelta(days=2), activity=activities[1])
        DateActivitySlotFactory.create(status='open', start=now() + timedelta(days=5), activity=activities[1])

        DateActivitySlotFactory.create(status='open', start=now() + timedelta(days=4), activity=activities[2])
        DateActivitySlotFactory.create(status='open', start=now() + timedelta(days=7), activity=activities[2])

        self.search({'upcoming': '1'}, 'date')

        self.assertEqual(
            [str(activity.pk) for activity in activities],
            [activity['id'] for activity in self.data['data']]
        )

    def test_filter_closed_segments(self):
        segment_type = SegmentTypeFactory.create(is_active=True, enable_search=True)
        open_segment = SegmentFactory.create(segment_type=segment_type, closed=False)
        closed_segment = SegmentFactory.create(segment_type=segment_type, closed=True)

        open = [
            DateActivityFactory.create(status='open'),
            CollectActivityFactory.create(status='open')
        ]
        for activity in open:
            activity.segments.add(open_segment)

        closed = [
            DateActivityFactory.create(status='open'),
            CollectActivityFactory.create(status='open')
        ]
        for activity in closed:
            activity.segments.add(closed_segment)

        self.search({})
        self.assertFound(open)

        user = BlueBottleUserFactory.create()
        user.segments.add(closed_segment)

        self.search({}, user=user)
        self.assertFound(open + closed)

        staff_user = BlueBottleUserFactory.create(is_staff=True)

        self.search({}, user=staff_user)
        self.assertFound(open + closed)

    def test_filter_type(self):
        matching = (
            DateActivityFactory.create_batch(3, status='open') +
            PeriodActivityFactory.create_batch(2, status='open')
        )
        funding = FundingFactory.create_batch(1, status='open')
        deed = DeedFactory.create_batch(3, status='open')
        collect = CollectActivityFactory.create_batch(4, status='open')

        self.search({'activity-type': 'time'})

        self.assertFacets(
            'activity-type',
            {
                'time': (None, len(matching)),
                'funding': (None, len(funding)),
                'collect': (None, len(collect)),
                'deed': (None, len(deed)),

            }
        )

        self.assertFound(matching)

    def test_filter_segment(self):
        segment_type = SegmentTypeFactory.create(is_active=True, enable_search=True)
        matching_segment, other_segment = SegmentFactory.create_batch(2, segment_type=segment_type)

        matching = [
            DateActivityFactory.create(status='open'),
            CollectActivityFactory.create(status='open')
        ]
        for activity in matching:
            activity.segments.add(matching_segment)

        other = [
            DateActivityFactory.create(status='open'),
            CollectActivityFactory.create(status='open')
        ]
        for activity in other:
            activity.segments.add(other_segment)

        self.search({f'segment.{segment_type.slug}': matching_segment.pk})

        self.assertFacets(
            f'segment.{segment_type.slug}',
            {
                f'{matching_segment.pk}': (matching_segment.name, len(matching)),
                f'{other_segment.pk}': (other_segment.name, len(other))
            }
        )
        self.assertFound(matching)

    def test_filter_theme(self):
        settings = InitiativePlatformSettings.objects.create()
        ActivitySearchFilter.objects.create(settings=settings, type="theme")

        matching_initiative, other_initiative = InitiativeFactory.create_batch(2, status='approved')

        matching = DeedFactory.create_batch(3, status="open", initiative=matching_initiative)
        other = DeedFactory.create_batch(2, status="open", initiative=other_initiative)

        self.search({
            'theme': matching_initiative.theme.pk,
        })

        self.assertFacets(
            'theme',
            {
                str(matching_initiative.theme.pk): (matching_initiative.theme.name, len(matching)),
                str(other_initiative.theme.pk): (other_initiative.theme.name, len(other))
            }
        )
        self.assertFound(matching)

    def test_filter_theme_no_matches(self):
        settings = InitiativePlatformSettings.objects.create()
        ActivitySearchFilter.objects.create(settings=settings, type="theme")
        ActivitySearchFilter.objects.create(settings=settings, type="country")

        matching_initiative, other_initiative = InitiativeFactory.create_batch(2, status='approved')

        DeedFactory.create_batch(3, status="open", initiative=matching_initiative)
        DeedFactory.create_batch(2, status="open", initiative=other_initiative)

        self.search({
            'theme': matching_initiative.theme.pk,
            'country': 'something-that-does-not-match'
        })

        self.assertFacets(
            'theme',
            {
                str(matching_initiative.theme.pk): (matching_initiative.theme.name, 0),
            }
        )
        self.assertFound([])

    def test_filter_theme_dutch(self):
        settings = InitiativePlatformSettings.objects.create()
        ActivitySearchFilter.objects.create(settings=settings, type="theme")

        matching_initiative, other_initiative = InitiativeFactory.create_batch(2, status='approved')

        matching = DeedFactory.create_batch(3, status="open", initiative=matching_initiative)
        other = DeedFactory.create_batch(2, status="open", initiative=other_initiative)

        self.search(
            {'theme': matching_initiative.theme.pk},
            headers={'HTTP_X_APPLICATION_LANGUAGE': 'nl'}
        )

        matching_theme_translation = matching_initiative.theme.translations.get(
            language_code='nl'
        )
        other_theme_translation = other_initiative.theme.translations.get(
            language_code='nl'
        )

        self.assertFacets(
            'theme',
            {
                str(matching_initiative.theme.pk): (matching_theme_translation.name, len(matching)),
                str(other_initiative.theme.pk): (other_theme_translation.name, len(other))
            }
        )
        self.assertFound(matching)

    def test_filter_initiative(self):

        initiator = BlueBottleUserFactory.create()
        activity_manager = BlueBottleUserFactory.create()
        draft_owner = BlueBottleUserFactory.create()
        random_user = BlueBottleUserFactory.create()

        initiative = InitiativeFactory.create(status='approved', owner=initiator)
        initiative.activity_managers.add(activity_manager)

        open = DeedFactory.create(status="open", initiative=initiative)
        draft = DeedFactory.create(status="draft", initiative=initiative, owner=draft_owner)
        DeedFactory.create(status="deleted", initiative=initiative)
        DeedFactory.create(status="open")

        self.search({'initiative.id': initiative.id}, user=initiator)
        self.assertFound([open, draft])

        self.search({'initiative.id': initiative.id}, user=activity_manager)
        self.assertFound([open, draft])

        self.search({'initiative.id': initiative.id}, user=draft_owner)
        self.assertFound([open, draft])

        self.search({'initiative.id': initiative.id}, user=random_user)
        self.assertFound([open])

        self.search({'initiative.id': initiative.id})
        self.assertFound([open])

    def test_filter_upcoming(self):
        matching = (
            PeriodActivityFactory.create_batch(2, status='open') +
            PeriodActivityFactory.create_batch(2, status='full')
        )
        other = PeriodActivityFactory.create_batch(2, status='succeeded')

        PeriodActivityFactory.create_batch(2, status='draft')
        PeriodActivityFactory.create_batch(2, status='needs_works')

        self.search({'upcoming': 1})

        self.assertFacets('upcoming', {0: ('No', len(other)), 1: ('Yes', len(matching))})
        self.assertFound(matching)

    def test_no_filter(self):
        matching = (
            PeriodActivityFactory.create_batch(2, status='open') +
            PeriodActivityFactory.create_batch(2, status='full') +
            PeriodActivityFactory.create_batch(2, status='full') +
            PeriodActivityFactory.create_batch(2, status='partially_funded')
        )
        PeriodActivityFactory.create_batch(2, status='draft')
        PeriodActivityFactory.create_batch(2, status='needs_works')
        PeriodActivityFactory.create_batch(2, status='needs_works')

        self.search({})

        self.assertFound(matching)

    def test_filter_team(self):
        InitiativePlatformSettings.objects.create(activity_search_filters=['team_activity'])

        matching = PeriodActivityFactory.create_batch(2, status="open", team_activity='teams')
        other = PeriodActivityFactory.create_batch(3, status="open", team_activity='individuals')

        self.search({'team_activity': 'teams'})

        self.assertFacets(
            'team_activity',
            {'teams': ('With your team', len(matching)), 'individuals': ('As an individual', len(other))}
        )
        self.assertFound(matching)

    def test_filter_online(self):
        matching = PeriodActivityFactory.create_batch(2, status="open", is_online=True)
        other = PeriodActivityFactory.create_batch(3, status="open", is_online=False)

        self.search({'is_online': '1'})

        self.assertFacets(
            'is_online',
            {1: ('Online/remote', len(matching)), 0: ('In-person', len(other))}
        )
        self.assertFound(matching)

    def test_filter_category(self):
        settings = InitiativePlatformSettings.objects.create()
        ActivitySearchFilter.objects.create(settings=settings, type="category")

        matching_category = CategoryFactory.create()
        other_category = CategoryFactory.create()

        matching = PeriodActivityFactory.create_batch(2, status='open')
        for activity in matching:
            activity.initiative.categories.add(matching_category)

        other = PeriodActivityFactory.create_batch(3, status='open')
        for activity in other:
            activity.initiative.categories.add(other_category)

        self.search({'category': matching_category.pk})

        self.assertFacets(
            'category',
            {
                str(matching_category.pk): (matching_category.title, len(matching)),
                str(other_category.pk): (other_category.title, len(other))
            }
        )
        self.assertFound(matching)

    def test_filter_skill(self):
        settings = InitiativePlatformSettings.objects.create()
        ActivitySearchFilter.objects.create(settings=settings, type="skill")

        matching_skill = SkillFactory.create()
        other_skill = SkillFactory.create()

        matching = PeriodActivityFactory.create_batch(
            2,
            expertise=matching_skill,
            status='open',
        )

        other = PeriodActivityFactory.create_batch(
            3,
            expertise=other_skill,
            status='open',
        )

        self.search({'skill': matching_skill.pk})

        self.assertFacets(
            'skill',
            {
                str(matching_skill.pk): (matching_skill.name, len(matching)),
                str(other_skill.pk): (other_skill.name, len(other))
            }
        )
        self.assertFound(matching)

    def test_filter_country(self):
        settings = InitiativePlatformSettings.objects.create()
        ActivitySearchFilter.objects.create(settings=settings, type="country")

        matching_country = CountryFactory.create()
        other_country = CountryFactory.create()

        matching = PeriodActivityFactory.create_batch(
            2,
            office_location=LocationFactory.create(country=matching_country),
            status='open',
        )

        other = PeriodActivityFactory.create_batch(
            3,
            office_location=LocationFactory.create(country=other_country),
            status='open',
        )

        self.search({'country': matching_country.pk})

        self.assertFacets(
            'country',
            {
                str(matching_country.pk): (matching_country.name, len(matching)),
                str(other_country.pk): (other_country.name, len(other))
            }
        )
        self.assertFound(matching)

    def test_more_country_facets(self):
        countries = CountryFactory.create_batch(12)
        matching = []
        for country in countries:
            location = GeolocationFactory.create(country=country)
            matching.append(PeriodActivityFactory.create(location=location, status='open'))

        self.search({})
        self.assertEqual(len(self.data['meta']['facets']['country']), 12)
        self.assertFound(matching)

    def test_filter_country_slots(self):
        settings = InitiativePlatformSettings.objects.create()
        ActivitySearchFilter.objects.create(settings=settings, type="country")

        matching_country = CountryFactory.create()
        other_country = CountryFactory.create()

        matching = DateActivityFactory.create_batch(
            2,
            status='open',
        )
        for activity in matching:
            DateActivitySlotFactory.create_batch(
                2,
                activity=activity,
                location=GeolocationFactory.create(country=matching_country)
            )

        other = DateActivityFactory.create_batch(
            3,
            status='open',
        )
        for activity in other:
            DateActivitySlotFactory.create_batch(
                2,
                activity=activity,
                location=GeolocationFactory.create(country=other_country)
            )

        self.search({'country': matching_country.pk})

        self.assertFacets(
            'country',
            {
                str(matching_country.pk): (matching_country.name, len(matching)),
                str(other_country.pk): (other_country.name, len(other))
            }
        )
        self.assertFound(matching)

    def test_filter_highlight(self):
        matching = PeriodActivityFactory.create_batch(
            2,
            highlight=True,
            status='open',
        )

        other = PeriodActivityFactory.create_batch(
            3,
            highlight=False,
            status='open',
        )
        self.search({'highlight': 'true'})

        self.assertFacets(
            'highlight',
            {1: (None, len(matching)), 0: (None, len(other))}
        )
        self.assertFound(matching)

    def test_filter_date(self):
        matching = [
            PeriodActivityFactory.create(status="open", start='2025-04-01', deadline='2025-04-02'),
            PeriodActivityFactory.create(status="open", start='2025-04-01', deadline='2025-04-03'),
            DeedFactory.create(status="open", start='2025-04-05', end='2025-04-07'),
            CollectActivityFactory.create(status="open", start='2025-04-05', end='2025-04-07'),
        ]

        PeriodActivityFactory.create(status="open", start='2025-05-01', deadline='2025-05-02')
        PeriodActivityFactory.create(status="open", start='2025-05-01', deadline='2025-05-03')
        DeedFactory.create(status="open", start='2025-05-05', end='2025-05-07')
        CollectActivityFactory.create(status="open", start='2025-05-05', end='2025-05-07')

        self.search({'date': '2025-04-01,2025-04-08'})

        self.assertFacets(
            'date', {}
        )

        self.assertFound(matching)

    def test_filter_past_dates(self):
        activity1 = DateActivityFactory.create(status="succeeded")
        activity2 = DateActivityFactory.create(status="succeeded")
        activity3 = DateActivityFactory.create(status="succeeded")
        activity4 = PeriodActivityFactory.create(status="succeeded", start='2022-04-15', deadline='2022-05-15')
        PeriodActivityFactory.create(status="succeeded", start='2022-03-01', deadline='2022-06-01')

        DateActivitySlotFactory.create(activity=activity1, start=datetime.datetime(2022, 5, 3, tzinfo=UTC))
        DateActivitySlotFactory.create(activity=activity1, start=datetime.datetime(2022, 5, 25, tzinfo=UTC))
        DateActivitySlotFactory.create(activity=activity1, start=datetime.datetime(2022, 6, 3, tzinfo=UTC))

        DateActivitySlotFactory.create(activity=activity2, start=datetime.datetime(2022, 5, 30, tzinfo=UTC))
        DateActivitySlotFactory.create(activity=activity2, start=datetime.datetime(2022, 5, 25, tzinfo=UTC))
        DateActivitySlotFactory.create(activity=activity2, start=datetime.datetime(2022, 4, 25, tzinfo=UTC))

        DateActivitySlotFactory.create(activity=activity3, start=datetime.datetime(2022, 6, 3, tzinfo=UTC))
        DateActivitySlotFactory.create(activity=activity3, start=datetime.datetime(2022, 4, 23, tzinfo=UTC))

        matching = [
            activity1, activity2, activity4
        ]

        self.search({'date': '2022-05-01,2022-05-31'})
        self.assertFound(matching)
        self.assertEqual(
            self.data['data'][0]['attributes']['end'],
            "2022-05-30T02:00:00+00:00"
        )
        self.assertEqual(
            self.data['data'][1]['attributes']['end'],
            "2022-05-25T02:00:00+00:00"
        )
        self.assertEqual(
            self.data['data'][2]['attributes']['end'],
            "2022-05-15"
        )

    def test_filter_distance(self):
        lat = 52.0
        lon = 10

        place = PlaceFactory.create(
            position=Point(lon, lat)
        )
        matching = [
            DateActivityFactory.create(status="open", slots=[]),
            DateActivityFactory.create(status="open", slots=[]),
            PeriodActivityFactory.create(
                status="open", location=GeolocationFactory.create(position=Point(lon + 0.1, lat + 0.1))
            ),
            PeriodActivityFactory.create(
                status="open", location=GeolocationFactory.create(position=Point(lon - 0.1, lat - 0.1))
            )
        ]

        DateActivitySlotFactory.create(
            activity=matching[0],
            location=GeolocationFactory.create(position=Point(lon + 0.05, lat + 0.05))
        )
        DateActivitySlotFactory.create(
            activity=matching[1],
            location=GeolocationFactory.create(position=Point(lon - 0.05, lat - 0.05))
        )

        PeriodActivityFactory.create(
            status="open",
            location=GeolocationFactory.create(position=Point(lon - 2, lat - 2))
        )
        PeriodActivityFactory.create(
            status="open",
            location=GeolocationFactory.create(position=Point(lon - 2, lat - 2))
        )
        DeedFactory.create()

        other = DateActivityFactory.create(slots=[])
        DateActivitySlotFactory.create(
            activity=other,
            location=GeolocationFactory.create(position=Point(lon + 2, lat + 2))
        )

        PeriodActivityFactory.create(
            status="open",
            is_online=True
        )

        self.search({'distance': '100km', 'is_online': '0'}, place=place.pk)

        self.assertFacets(
            'distance', {}
        )

        self.assertFound(matching)

    def test_filter_distance_with_online(self):
        amsterdam = Point(4.922114, 52.362438)
        leiden = Point(4.491056, 52.166758)
        lyutidol = Point(23.676222, 43.068555)

        place = PlaceFactory(position=leiden)
        matching = [
            DateActivityFactory.create(status="open", slots=[]),
            DateActivityFactory.create(status="open", slots=[]),
            PeriodActivityFactory.create(
                status="open", location=GeolocationFactory.create(position=leiden),
            ),
            PeriodActivityFactory.create(
                status="open", location=GeolocationFactory.create(position=amsterdam),
            ),
            PeriodActivityFactory.create(
                status="open", is_online=True,
            )

        ]

        DateActivitySlotFactory.create(
            activity=matching[0],
            location=GeolocationFactory.create(position=leiden)
        )
        DateActivitySlotFactory.create(
            activity=matching[1],
            location=GeolocationFactory.create(position=amsterdam)
        )

        PeriodActivityFactory.create(
            status="open", location=GeolocationFactory.create(position=lyutidol)
        )
        PeriodActivityFactory.create(
            status="open", location=GeolocationFactory.create(position=lyutidol)
        )
        DeedFactory.create(status="open")

        other = DateActivityFactory.create(status="open", slots=[])
        DateActivitySlotFactory.create(
            activity=other,
            location=GeolocationFactory.create(position=lyutidol)
        )

        self.search({'distance': '100km'}, place=place.pk)

        self.assertFacets(
            'distance', {}
        )

        self.assertFound(matching)

    def test_filter_office_restriction(self):
        office = LocationFactory.create(subregion=OfficeSubRegionFactory.create())
        within_sub_region = LocationFactory.create(subregion=office.subregion)
        within_region = LocationFactory.create(
            subregion=OfficeSubRegionFactory(region=office.subregion.region)
        )

        matching = [
            PeriodActivityFactory.create(
                status="open", office_location=office, office_restriction='office'
            ),
            PeriodActivityFactory.create(
                status="open", office_location=within_region, office_restriction='office_region'
            ),
            PeriodActivityFactory.create(
                status="open", office_location=within_sub_region, office_restriction='office_subregion'
            ),
            PeriodActivityFactory.create(
                status="open", office_location=LocationFactory.create(), office_restriction='all'
            ),
        ]

        PeriodActivityFactory.create(
            status="open", office_location=LocationFactory.create(), office_restriction='office'
        )
        PeriodActivityFactory.create(
            status="open", office_location=within_region, office_restriction='office'
        )
        PeriodActivityFactory.create(
            status="open", office_location=within_region, office_restriction='office_subregion'
        )

        user = BlueBottleUserFactory.create(location=office)

        self.search({'office_restriction': '1'}, user=user)

        self.assertFacets(
            'distance', {}
        )

        self.assertFound(matching)


class ActivityRelatedImageAPITestCase(BluebottleTestCase):
    def setUp(self):
        super(ActivityRelatedImageAPITestCase, self).setUp()
        self.client = JSONAPITestClient()
        self.owner = BlueBottleUserFactory.create()
        self.funding = FundingFactory.create(
            owner=self.owner,
        )
        self.related_image_url = reverse('related-activity-image-list')

        file_path = './bluebottle/files/tests/files/test-image.png'

        with open(file_path, 'rb') as test_file:
            response = self.client.post(
                reverse('image-list'),
                test_file.read(),
                content_type="image/png",
                format=None,
                HTTP_CONTENT_DISPOSITION='attachment; filename="some_file.jpg"',
                user=self.owner
            )

        self.file_data = json.loads(response.content)

    def test_create(self):
        data = {
            'data': {
                'type': 'related-activity-images',
                'relationships': {
                    'image': {
                        'data': {
                            'type': 'images',
                            'id': self.file_data['data']['id']
                        }
                    },
                    'resource': {
                        'data': {
                            'type': 'activities/fundings',
                            'id': self.funding.pk,
                        }
                    }
                }
            }
        }
        response = self.client.post(
            self.related_image_url,
            data=json.dumps(data),
            user=self.owner
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        self.assertEqual(
            response.json()['included'][1]['attributes']['links']['large'].split('?')[0],
            u'/api/activities/{}/related-image/600'.format(response.json()['data']['id'])
        )

    def test_create_non_owner(self):
        data = {
            'data': {
                'type': 'related-activity-images',
                'relationships': {
                    'image': {
                        'data': {
                            'type': 'images',
                            'id': self.file_data['data']['id']
                        }
                    },
                    'resource': {
                        'data': {
                            'type': 'activities/fundings',
                            'id': self.funding.pk,
                        }
                    }
                }
            }
        }
        response = self.client.post(
            self.related_image_url,
            data=json.dumps(data),
            user=BlueBottleUserFactory.create()
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


class ContributorListAPITestCase(BluebottleTestCase):
    def setUp(self):
        super(ContributorListAPITestCase, self).setUp()
        self.client = JSONAPITestClient()
        self.user = BlueBottleUserFactory.create()

        participants = DateParticipantFactory.create_batch(2, user=self.user)
        for participant in participants:
            slot = DateActivitySlotFactory.create(activity=participant.activity)
            slot.slot_participants.all()[0].states.remove(save=True)

        PeriodParticipantFactory.create_batch(2, user=self.user)
        DonorFactory.create_batch(2, user=self.user, status='succeeded')
        DonorFactory.create_batch(2, user=self.user, status='new')
        DeedParticipantFactory.create_batch(2, user=self.user)
        CollectContributorFactory.create_batch(2, user=self.user)

        DateParticipantFactory.create()
        PeriodParticipantFactory.create()
        DonorFactory.create()
        DeedParticipantFactory.create()
        CollectContributorFactory.create()

        self.url = reverse('contributor-list')

    def test_get(self):
        response = self.client.get(
            self.url,
            user=self.user
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()

        self.assertEqual(len(data['data']), 10)

        for contributor in data['data']:
            self.assertTrue(
                contributor['type'] in (
                    'contributors/time-based/date-participants',
                    'contributors/time-based/period-participants',
                    'contributors/collect/contributors',
                    'contributors/deeds/participant',
                    'contributors/donations',
                )
            )
            self.assertTrue(contributor['type'])
            self.assertTrue(
                contributor['relationships']['activity']['data']['type'] in (
                    'activities/fundings',
                    'activities/deeds',
                    'activities/time-based/dates',
                    'activities/time-based/periods',
                    'activities/collects'
                )
            )

            if contributor['type'] == 'contributors/time-based/date-participants':
                self.assertEqual(contributor['attributes']['total-duration'], '02:00:00')

        self.assertEqual(
            len([
                resource for resource in data['included']
                if resource['type'] == 'activities/time-based/periods'
            ]),
            2
        )

        self.assertEqual(
            len([
                resource for resource in data['included']
                if resource['type'] == 'activities/time-based/dates'
            ]),
            2
        )

        self.assertEqual(
            len([
                resource for resource in data['included']
                if resource['type'] == 'activities/deeds'
            ]),
            2
        )

        self.assertEqual(
            len([
                resource for resource in data['included']
                if resource['type'] == 'activities/fundings'
            ]),
            2
        )

    def test_get_anonymous(self):
        response = self.client.get(
            self.url
        )

        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_get_other_user(self):
        response = self.client.get(
            self.url,
            user=BlueBottleUserFactory.create()
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()

        self.assertEqual(len(data['data']), 0)


@override_settings(
    ELASTICSEARCH_DSL_AUTOSYNC=True,
    ELASTICSEARCH_DSL_AUTO_REFRESH=True
)
@tag('elasticsearch')
class ActivityAPIAnonymizationTestCase(ESTestCase, BluebottleTestCase):
    anonymous_resource = {
        'id': 'anonymous',
        'type': 'members',
        'attributes': {
            'is-anonymous': True
        }
    }

    def setUp(self):
        super(ActivityAPIAnonymizationTestCase, self).setUp()
        self.member_settings = MemberPlatformSettings.load()

        self.client = JSONAPITestClient()
        self.owner = BlueBottleUserFactory.create()

    def test_activity_over_max_age(self):
        self.member_settings.anonymization_age = 300
        self.member_settings.save()

        activity = DateActivityFactory.create(
            created=now() - timedelta(days=400),
            status='open'
        )

        data = self.client.get(
            reverse('date-detail', args=(activity.id,))
        ).json()

        self.assertEqual(
            data['data']['relationships']['owner']['data']['id'], 'anonymous'
        )

        self.assertTrue(self.anonymous_resource in data['included'])

    def test_activity_not_over_max_age(self):
        self.member_settings.anonymization_age = 300
        self.member_settings.save()

        activity = DateActivityFactory.create(
            created=now() - timedelta(days=200),
            status='open'
        )

        data = self.client.get(
            reverse('date-detail', args=(activity.id,))
        ).json()

        self.assertEqual(
            data['data']['relationships']['owner']['data']['id'], str(activity.owner.pk)
        )

        self.assertTrue(self.anonymous_resource not in data['included'])

    def test_initiative_over_max_age(self):
        self.member_settings.anonymization_age = 300
        self.member_settings.save()

        initiative = InitiativeFactory.create(
            status='open',
            promoter=BlueBottleUserFactory.create(),
            reviewer=BlueBottleUserFactory.create(),
        )

        initiative.created = now() - timedelta(days=400)
        initiative.save()

        DateActivityFactory.create(
            initiative=initiative,
            created=now() - timedelta(days=400),
            status='open'
        )
        data = self.client.get(
            reverse('initiative-detail', args=(initiative.id,))
        ).json()

        self.assertEqual(
            data['data']['relationships']['owner']['data']['id'], 'anonymous'
        )

        self.assertEqual(
            data['data']['relationships']['activity-managers']['data'][0]['id'], 'anonymous'
        )

        self.assertEqual(
            data['data']['relationships']['reviewer']['data']['id'], 'anonymous'
        )

    def test_initiative_not_over_max_age(self):
        self.member_settings.anonymization_age = 300
        self.member_settings.save()

        initiative = InitiativeFactory.create(
            status='open',
            promoter=BlueBottleUserFactory.create(),
            reviewer=BlueBottleUserFactory.create(),
        )

        initiative.created = now() - timedelta(days=200)
        initiative.save()

        DateActivityFactory.create(
            initiative=initiative,
            status='open'
        )
        data = self.client.get(
            reverse('initiative-detail', args=(initiative.id,))
        ).json()

        self.assertEqual(
            data['data']['relationships']['owner']['data']['id'], str(initiative.owner.pk)
        )

        self.assertEqual(
            data['data']['relationships']['activity-managers']['data'][0]['id'],
            str(initiative.activity_managers.first().pk)
        )

        self.assertEqual(
            data['data']['relationships']['reviewer']['data']['id'], str(initiative.reviewer.pk)
        )

    def test_participants_over_max_age(self):
        self.member_settings.anonymization_age = 300
        self.member_settings.save()

        activity = DateActivityFactory.create(
            created=now() - timedelta(days=400),
            status='open'
        )

        activity_data = self.client.get(
            reverse('date-detail', args=(activity.id,))
        ).json()

        DateParticipantFactory.create(
            activity=activity,
            created=now() - timedelta(days=350),
        )
        new_participant = DateParticipantFactory.create(
            activity=activity
        )

        data = self.client.get(
            activity_data['data']['relationships']['contributors']['links']['related'],
            user=self.owner
        ).json()
        self.assertEqual(
            data['data'][1]['relationships']['user']['data']['id'],
            'anonymous'
        )
        self.assertEqual(
            data['data'][0]['relationships']['user']['data']['id'],
            str(new_participant.user.pk)
        )


class TeamListViewAPITestCase(APITestCase):
    serializer = TeamSerializer

    def setUp(self):
        super().setUp()

        self.activity = PeriodActivityFactory.create(status='open')

        self.approved_teams = TeamFactory.create_batch(5, activity=self.activity)
        for team in self.approved_teams:
            PeriodParticipantFactory.create(activity=self.activity, team=team, user=team.owner)

        for team in self.approved_teams[:2]:
            TeamSlotFactory.create(activity=self.activity, team=team, start=now() + timedelta(days=5))

        TeamSlotFactory.create(
            activity=self.activity, team=self.approved_teams[2], start=now() - timedelta(days=5)
        )

        self.cancelled_teams = TeamFactory.create_batch(
            5, activity=self.activity, status='cancelled'
        )
        for team in self.cancelled_teams:
            PeriodParticipantFactory.create(activity=self.activity, team=team, user=team.owner)
            PeriodParticipantFactory.create(activity=self.activity, team=team)

        self.url = "{}?filter[activity_id]={}".format(
            reverse('team-list'),
            self.activity.pk
        )

        settings = InitiativePlatformSettings.objects.get()
        settings.team_activities = True
        settings.enable_participant_exports = True
        settings.save()

    def test_get_activity_owner(self):
        self.perform_get(user=self.activity.owner)

        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(len(self.approved_teams) + len(self.cancelled_teams))
        self.assertObjectList(self.approved_teams)
        self.assertRelationship('owner')

        self.assertMeta('status')
        self.assertMeta('transitions')
        for resource in self.response.json()['data']:
            self.assertTrue(resource['meta']['participants-export-url'] is not None)
        team_ids = [t["id"] for t in self.response.json()["data"]]
        self.assertEqual(
            len(team_ids),
            len(set(team_ids)),
            'We should have a unique list of team ids'
        )

    def test_get_filtered_status(self):
        new_teams = TeamFactory.create_batch(2, activity=self.activity, status='new')
        self.perform_get(user=self.activity.owner, query={'filter[status]': 'new'})

        self.assertStatus(status.HTTP_200_OK)
        for resource in self.response.json()['data']:
            self.assertTrue(
                resource['id'] in [str(team.pk) for team in new_teams]
            )

    def test_get_filtered_has_no_slot(self):
        self.perform_get(user=self.activity.owner, query={'filter[has_slot]': 'false'})

        self.assertStatus(status.HTTP_200_OK)
        for resource in self.response.json()['data']:
            self.assertIsNone(resource['relationships']['slot']['data'])

    def test_get_filtered_future(self):
        self.perform_get(user=self.activity.owner, query={'filter[start]': 'future'})

        self.assertStatus(status.HTTP_200_OK)
        for resource in self.response.json()['data']:
            self.assertTrue(
                resource['id'] in [str(team.pk) for team in self.approved_teams[:2]]
            )

    def test_get_filtered_passed(self):
        self.perform_get(user=self.activity.owner, query={'filter[start]': 'passed'})

        self.assertStatus(status.HTTP_200_OK)
        for resource in self.response.json()['data']:
            self.assertEqual(
                resource['id'], str(self.approved_teams[2].pk)
            )

    def test_get_cancelled_team_captain(self):
        team = self.cancelled_teams[0]
        self.perform_get(user=team.owner)

        self.assertStatus(status.HTTP_200_OK)

        self.assertTotal(len(self.approved_teams) + 1)
        self.assertObjectList(self.approved_teams + [team])
        self.assertRelationship('owner')

        self.assertEqual(
            self.response.json()['data'][0]['relationships']['owner']['data']['id'],
            str(team.owner.pk)
        )

    def test_get_team_captain(self):
        team = self.approved_teams[0]
        self.perform_get(user=team.owner)

        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(len(self.approved_teams))
        self.assertObjectList(self.approved_teams)
        self.assertRelationship('owner')

        self.assertEqual(
            self.response.json()['data'][0]['relationships']['owner']['data']['id'],
            str(team.owner.pk)
        )

        for resource in self.response.json()['data']:
            if resource['relationships']['owner']['data']['id'] == str(team.owner.pk):
                self.assertTrue(resource['meta']['participants-export-url'] is not None)
            else:
                self.assertTrue(resource['meta']['participants-export-url'] is None)

    def test_get_anonymous(self):
        self.perform_get()

        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(len(self.approved_teams))
        self.assertObjectList(self.approved_teams)
        self.assertRelationship('owner')
        for resource in self.response.json()['data']:
            self.assertTrue(resource['meta']['participants-export-url'] is None)

    def test_pagination(self):
        extra_teams = TeamFactory.create_batch(
            10, activity=self.activity
        )
        self.perform_get()
        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(len(self.approved_teams) + len(extra_teams))
        self.assertSize(8)
        self.assertPages(2)

    def test_other_user_anonymous(self):
        self.perform_get(BlueBottleUserFactory.create())

        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(len(self.approved_teams))
        self.assertObjectList(self.approved_teams)
        self.assertRelationship('owner')

    def test_get_anonymous_closed_site(self):
        with self.closed_site():
            self.perform_get()

        self.assertStatus(status.HTTP_401_UNAUTHORIZED)

    def test_get_user_closed_site(self):
        with self.closed_site():
            self.perform_get(BlueBottleUserFactory.create())

        self.assertStatus(status.HTTP_200_OK)


class TeamTransitionListViewAPITestCase(APITestCase):
    url = reverse('team-transition-list')
    serializer = TeamTransitionSerializer

    def setUp(self):
        super().setUp()

        self.team = TeamFactory.create()

        self.defaults = {
            'resource': self.team,
            'transition': 'cancel',
        }

        self.fields = ['resource', 'transition', ]

    def test_cancel_owner(self):
        self.perform_create(user=self.team.owner)
        self.assertStatus(status.HTTP_400_BAD_REQUEST)

        self.team.refresh_from_db()
        self.assertEqual(self.team.status, 'open')

    def test_cancel_activity_manager(self):
        self.perform_create(user=self.team.activity.owner)

        self.assertStatus(status.HTTP_201_CREATED)
        self.assertIncluded('resource', self.team)

        self.team.refresh_from_db()
        self.assertEqual(self.team.status, 'cancelled')

    def test_cancel_other_user(self):
        self.perform_create(user=BlueBottleUserFactory.create())
        self.assertStatus(status.HTTP_400_BAD_REQUEST)

        self.team.refresh_from_db()
        self.assertEqual(self.team.status, 'open')

    def test_cancel_no_user(self):
        self.perform_create()
        self.assertStatus(status.HTTP_400_BAD_REQUEST)

        self.team.refresh_from_db()
        self.assertEqual(self.team.status, 'open')

    def test_withdraw_owner(self):
        self.defaults['transition'] = 'withdraw'

        self.perform_create(user=self.team.owner)

        self.assertStatus(status.HTTP_201_CREATED)
        self.assertIncluded('resource', self.team)

        self.team.refresh_from_db()
        self.assertEqual(self.team.status, 'withdrawn')

    def test_withdraw_activity_manager(self):
        self.defaults['transition'] = 'withdraw'

        self.perform_create(user=self.team.activity.owner)

        self.assertStatus(status.HTTP_400_BAD_REQUEST)
        self.team.refresh_from_db()

        self.assertEqual(self.team.status, 'open')

    def test_withdraw_other_user(self):
        self.defaults['transition'] = 'withdraw'

        self.perform_create(user=BlueBottleUserFactory.create())
        self.assertStatus(status.HTTP_400_BAD_REQUEST)

        self.team.refresh_from_db()
        self.assertEqual(self.team.status, 'open')

    def test_withdraw_no_user(self):
        self.defaults['transition'] = 'withdraw'

        self.perform_create()
        self.assertStatus(status.HTTP_400_BAD_REQUEST)

        self.team.refresh_from_db()
        self.assertEqual(self.team.status, 'open')


class InviteDetailViewAPITestCase(APITestCase):
    serializer = InviteSerializer

    def setUp(self):
        super().setUp()
        activity = PeriodActivityFactory.create(status='open', team_activity='teams')
        self.contributor = PeriodParticipantFactory.create(activity=activity)

        self.url = reverse('invite-detail', args=(self.contributor.invite.pk,))

    def test_get_anonymous(self):
        self.perform_get()
        self.assertStatus(status.HTTP_200_OK)

        self.assertIncluded('team', self.contributor.team)
        self.assertIncluded('team.owner', self.contributor.team.owner)

    def test_get_anonymous_closed_site(self):
        with self.closed_site():
            self.perform_get()

        self.assertStatus(status.HTTP_401_UNAUTHORIZED)

    def test_get_anonymous_user(self):
        with self.closed_site():
            self.perform_get(user=BlueBottleUserFactory.create())

        self.assertStatus(status.HTTP_200_OK)


class TeamMemberExportViewAPITestCase(APITestCase):
    def setUp(self):
        super().setUp()

        settings = InitiativePlatformSettings.load()
        settings.team_activities = True
        settings.enable_participant_exports = True
        settings.save()

        self.activity = PeriodActivityFactory.create(team_activity='teams')

        self.team_captain = PeriodParticipantFactory(activity=self.activity)

        self.team_members = PeriodParticipantFactory.create_batch(
            3,
            activity=self.activity,
            accepted_invite=self.team_captain.invite
        )

        self.non_team_members = PeriodParticipantFactory.create_batch(
            3,
            activity=self.activity,
        )

        self.url = "{}?filter[activity_id]={}".format(reverse('team-list'), self.activity.pk)

    @property
    def export_url(self):
        for team in self.response.json()['data']:
            if team['id'] == str(self.team_captain.team.pk) and team['meta']['participants-export-url']:
                return team['meta']['participants-export-url']['url']

    def test_get_owner(self):
        self.perform_get(user=self.activity.owner)
        self.assertStatus(status.HTTP_200_OK)
        self.assertTrue(self.export_url)
        response = self.client.get(self.export_url)

        sheet = load_workbook(filename=io.BytesIO(response.content)).get_active_sheet()
        rows = list(sheet.values)

        self.assertEqual(
            rows[0],
            ('Email', 'Name', 'Registration Date', 'Status', 'Team Captain')
        )

        self.assertEqual(len(rows), 5)

        for team_member in self.team_members:
            self.assertTrue(team_member.user.email in [row[0] for row in rows[1:]])

        self.assertEqual(
            [
                row[4] for row in rows
                if row[0] == self.team_captain.user.email
            ][0],
            True
        )

    def test_team_captain(self):
        self.perform_get(user=self.team_captain.user)
        self.assertStatus(status.HTTP_200_OK)
        self.assertTrue(self.export_url)
        response = self.client.get(self.export_url)
        sheet = load_workbook(filename=io.BytesIO(response.content)).get_active_sheet()
        rows = list(sheet.values)

        self.assertEqual(
            rows[0],
            ('Email', 'Name', 'Registration Date', 'Status', 'Team Captain')
        )

    def test_get_owner_incorrect_hash(self):
        self.perform_get(user=self.activity.owner)
        self.assertStatus(status.HTTP_200_OK)
        response = self.client.get(self.export_url + 'test')
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_get_contributor(self):
        self.perform_get(user=self.team_members[0].user)
        self.assertIsNone(self.export_url)

    def test_get_other_user(self):
        self.perform_get(user=BlueBottleUserFactory.create())
        self.assertIsNone(self.export_url)

    def test_get_no_user(self):
        self.perform_get()
        self.assertIsNone(self.export_url)


class TeamMemberListViewAPITestCase(APITestCase):
    serializer = PeriodParticipantSerializer

    def setUp(self):
        super().setUp()

        settings = InitiativePlatformSettings.objects.get()
        settings.team_activities = True
        settings.save()

        self.activity = PeriodActivityFactory.create(status='open', team_activity='teams')

        self.team_captain = PeriodParticipantFactory.create(
            activity=self.activity
        )
        self.team = self.team_captain.team

        self.accepted_members = PeriodParticipantFactory.create_batch(
            3,
            activity=self.activity,
            accepted_invite=self.team_captain.invite
        )
        self.withdrawn_members = PeriodParticipantFactory.create_batch(
            3,
            activity=self.activity,
            accepted_invite=self.team_captain.invite
        )

        for member in self.withdrawn_members:
            member.states.withdraw(save=True)

        self.url = reverse('team-members', args=(self.team.pk,))

    def test_get_activity_owner(self):
        self.perform_get(user=self.activity.owner)

        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(len(self.accepted_members) + len(self.withdrawn_members) + 1)
        self.assertRelationship('user')

        self.assertAttribute('status')
        self.assertMeta('transitions')

    def test_get_team_captain(self):
        self.perform_get(user=self.team.owner)

        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(len(self.accepted_members) + len(self.withdrawn_members) + 1)
        ids = [m.id for m in self.accepted_members] + [m.id for m in self.withdrawn_members]

        self.assertEqual(len(set(ids)), 6)
        self.assertObjectList(self.accepted_members + self.withdrawn_members + [self.team_captain])
        self.assertRelationship('user')

        self.assertAttribute('status')
        self.assertMeta('transitions')

        self.assertTrue(
            str(self.team.owner.pk) in
            [m['relationships']['user']['data']['id'] for m in self.response.json()['data']]
        )

    def test_get_team_member(self):
        self.perform_get(user=self.accepted_members[0].user)

        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(len(self.accepted_members) + 1)

        self.assertObjectList(self.accepted_members + [self.team_captain])
        self.assertRelationship('user')

        self.assertAttribute('status')
        self.assertMeta('transitions')

    def test_get_other_user(self):
        self.perform_get(user=BlueBottleUserFactory.create())

        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(len(self.accepted_members) + 1)

        self.assertObjectList(self.accepted_members + [self.team_captain])
        self.assertRelationship('user')

        self.assertAttribute('status')
        self.assertMeta('transitions')

    def test_get_anonymous_closed_site(self):
        with self.closed_site():
            self.perform_get()

        self.assertStatus(status.HTTP_401_UNAUTHORIZED)


class ActivityLocationAPITestCase(APITestCase):
    model = Activity

    def setUp(self):
        CollectActivityFactory.create(status='succeeded')
        PeriodActivityFactory.create(status='succeeded')

        date_activity = DateActivityFactory.create(status="succeeded")
        date_activity.slots.add(DateActivitySlotFactory.create(activity=date_activity))

        self.url = reverse('activity-location-list')

    def test_get(self):
        self.perform_get()
        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(4)
        self.assertAttribute('position')
        self.assertRelationship('activity')

    def test_get_closed_platform(self):
        with self.closed_site():
            self.perform_get()
        self.assertStatus(status.HTTP_401_UNAUTHORIZED)

    def test_get_closed_platform_logged_in(self):
        with self.closed_site():
            self.perform_get(user=BlueBottleUserFactory.create())
        self.assertStatus(status.HTTP_200_OK)
