import os
from argparse import ArgumentTypeError
from datetime import datetime
from mock import patch
from openpyxl import load_workbook
import pytz

from django.conf import settings
from django.core.management import call_command
from django.core.urlresolvers import reverse
from django.test import SimpleTestCase

from bluebottle.analytics.management.commands.export_engagement_metrics import (
    Command as EngagementCommand
)
from bluebottle.analytics.management.commands.export_participation_metrics import (
    Command as ParticipationCommand,
)
from bluebottle.analytics.management.commands.export_analytics_data import (
    Command as AnalyticsCommand,
)
from bluebottle.bb_projects.models import ProjectPhase
from bluebottle.tasks.models import TaskMember
from bluebottle.test.factory_models.accounts import BlueBottleUserFactory
from bluebottle.test.factory_models.donations import DonationFactory
from bluebottle.test.factory_models.fundraisers import FundraiserFactory
from bluebottle.test.factory_models.orders import OrderFactory
from bluebottle.test.factory_models.projects import ProjectFactory
from bluebottle.test.factory_models.tasks import TaskFactory, TaskMemberFactory
from bluebottle.test.factory_models.votes import VoteFactory
from bluebottle.test.factory_models.wallposts import TextWallpostFactory
from bluebottle.test.utils import BluebottleTestCase

from .common import FakeInfluxDBClient


fake_client = FakeInfluxDBClient()


class TestEngagementMetricsUnit(SimpleTestCase):

    def test_validate_date(self):
        with self.assertRaises(ArgumentTypeError):
            EngagementCommand._validate_date('2017-13-13')
        self.assertEquals('2017-01-01', EngagementCommand._validate_date('2017-01-01'))

    def test_get_engagement_score(self):
        entry = {'comments': 1,
                 'votes': 1,
                 'donations': 1,
                 'tasks': 1,
                 'fundraisers': 1,
                 'projects': 1}

        self.assertEquals(30, EngagementCommand.get_engagement_score(entry))

    def test_get_engagement_rating(self):
        self.assertEquals('not engaged', EngagementCommand.get_engagement_rating(0))
        self.assertEquals('little engaged', EngagementCommand.get_engagement_rating(1))
        self.assertEquals('little engaged', EngagementCommand.get_engagement_rating(4))
        self.assertEquals('engaged', EngagementCommand.get_engagement_rating(5))
        self.assertEquals('engaged', EngagementCommand.get_engagement_rating(8))
        self.assertEquals('very engaged', EngagementCommand.get_engagement_rating(9))
        self.assertEquals('invalid engagement score: test', EngagementCommand.get_engagement_rating('test'))

    @patch('bluebottle.analytics.management.commands.export_engagement_metrics.datetime')
    def test_get_xls_file_name(self, mock_datetime):
        now = datetime(2017, 1, 1)
        mock_datetime.now.return_value = now
        start_date = datetime(2016, 1, 1)
        end_date = datetime(2016, 12, 31)
        self.assertEquals(EngagementCommand.get_xls_file_name(start_date, end_date),
                          'engagement_report_20160101_20161231_generated_20170101-000000.xlsx')


class TestEngagementMetricsXls(BluebottleTestCase):

    def setUp(self):
        super(TestEngagementMetricsXls, self).setUp()
        self.init_projects()

        self.year = datetime.now().year

        # Project Phases
        done_complete = ProjectPhase.objects.get(slug="done-complete")
        done_incomplete = ProjectPhase.objects.get(slug="done-incomplete")

        # Users
        user1 = BlueBottleUserFactory.create()

        # Projects
        project1 = ProjectFactory.create(owner=user1, status=done_complete)
        ProjectFactory.create(owner=user1, status=done_incomplete)

        # Wallposts
        TextWallpostFactory.create(content_object=project1,
                                   author=user1,
                                   editor=user1,
                                   text="test1",
                                   email_followers=False)

        # Votes
        VoteFactory(project=project1, voter=user1)

        # Fundraisers
        fundraiser = FundraiserFactory(project=project1, owner=user1)

        # Donations
        order1 = OrderFactory.create(user=user1)
        DonationFactory(order=order1, fundraiser=fundraiser, project=project1)
        order1.locked()
        order1.save()
        order1.success()
        order1.save()

        order2 = OrderFactory.create(user=None)
        donation2 = DonationFactory(order=order2, fundraiser=fundraiser, project=project1)
        donation2.anonymous = True
        order2.locked()
        order2.save()
        order2.success()
        order2.save()

        # Tasks
        task = TaskFactory.create(author=user1, project=project1, people_needed=2, status='realized')
        task_member = TaskMemberFactory.create(time_spent=10, member=user1, task=task,
                                               status=TaskMember.TaskMemberStatuses.applied)
        task_member.status = TaskMember.TaskMemberStatuses.realized
        task_member.save()

        # Simulate user Login
        jwt_token = user1.get_jwt_token()
        task_member_url = reverse('task-member-detail', kwargs={'pk': task_member.id})
        self.client.get(task_member_url, token="JWT {0}".format(jwt_token))

        # xls export
        self.xls_file_name = 'test.xlsx'
        self.xls_file_path = os.path.join(settings.PROJECT_ROOT, self.xls_file_name)
        self.command = EngagementCommand()

    def test_xls_generation(self):
        with patch.object(self.command, 'get_xls_file_name', return_value=self.xls_file_name):
            call_command(self.command, '--start', '{}-01-01'.format(self.year),
                         '--end', '{}-12-31'.format(self.year), '--export-to', 'xls')
            self.assertTrue(os.path.isfile(self.xls_file_path), True)

            workbook = load_workbook(filename=self.xls_file_path, read_only=True)
            worksheet = workbook['Engagement Aggregated Data']

            # Test Headers
            self.assertEqual(worksheet.cell(row=1, column=1).value, 'organisation')
            self.assertEqual(worksheet.cell(row=1, column=2).value, 'total no. of platforms')
            self.assertEqual(worksheet.cell(row=1, column=3).value, 'total members')
            self.assertEqual(worksheet.cell(row=1, column=4).value, 'not engaged members (engagement score: 0)')
            self.assertEqual(worksheet.cell(row=1, column=5).value, 'little engaged members (engagement score: 1-3)')
            self.assertEqual(worksheet.cell(row=1, column=6).value, 'engaged members (engagement score: 4-8)')
            self.assertEqual(worksheet.cell(row=1, column=7).value, 'very engaged members (engagement score: >8)')
            self.assertEqual(worksheet.cell(row=1, column=8).value, 'total engaged members (engagement score: >4)')
            self.assertEqual(worksheet.cell(row=1, column=9).value, '% total engaged members (engagement score: >4)')
            self.assertEqual(worksheet.cell(row=1, column=10).value, 'Projects Realised')
            self.assertEqual(worksheet.cell(row=1, column=11).value, 'Projects Done')
            self.assertEqual(worksheet.cell(row=1, column=12).value, 'Guest Donations')

            # Test Tenant test2
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'test2')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 1, msg='no. of platforms')
            self.assertEqual(worksheet.cell(row=2, column=3).value, 1, msg='total members')
            self.assertEqual(worksheet.cell(row=2, column=4).value, 1, msg='not engaged members')
            self.assertEqual(worksheet.cell(row=2, column=5).value, 0, msg='little engaged members')
            self.assertEqual(worksheet.cell(row=2, column=6).value, 0, msg='engaged members')
            self.assertEqual(worksheet.cell(row=2, column=7).value, 0, msg='very engaged members')
            self.assertEqual(worksheet.cell(row=2, column=8).value, 0, msg='total engaged members')
            self.assertEqual(worksheet.cell(row=2, column=9).value, 0, msg='% total engaged members')
            self.assertEqual(worksheet.cell(row=2, column=10).value, 0, msg='projects realised')
            self.assertEqual(worksheet.cell(row=2, column=11).value, 0, msg='projects done')
            self.assertEqual(worksheet.cell(row=2, column=12).value, 0, msg='guest donations')

            # Test Tenant test
            self.assertEqual(worksheet.cell(row=3, column=1).value, 'test')
            self.assertEqual(worksheet.cell(row=3, column=2).value, 1, msg='no. of platforms')
            self.assertEqual(worksheet.cell(row=3, column=3).value, 2, msg='total members')
            self.assertEqual(worksheet.cell(row=3, column=4).value, 1, msg='not engaged members')
            self.assertEqual(worksheet.cell(row=3, column=5).value, 0, msg='little engaged members')
            self.assertEqual(worksheet.cell(row=3, column=6).value, 0, msg='engaged members')
            self.assertEqual(worksheet.cell(row=3, column=7).value, 1, msg='very engaged members')
            self.assertEqual(worksheet.cell(row=3, column=8).value, 1, msg='total engaged members')
            self.assertEqual(worksheet.cell(row=3, column=9).value, 50, msg='% total engaged members')
            self.assertEqual(worksheet.cell(row=3, column=10).value, 1, msg='projects realised')
            self.assertEqual(worksheet.cell(row=3, column=11).value, 1, msg='projects done')
            self.assertEqual(worksheet.cell(row=3, column=12).value, 1, msg='guest donations')

    def tearDown(self):
        os.remove(self.xls_file_path)


class TestParticipationXls(BluebottleTestCase):

    def setUp(self):
        super(TestParticipationXls, self).setUp()
        self.init_projects()
        self.year = datetime.now().year

        # Project Phases
        done_complete = ProjectPhase.objects.get(slug="done-complete")
        done_incomplete = ProjectPhase.objects.get(slug="done-incomplete")

        # Users
        self.users = BlueBottleUserFactory.create_batch(200)

        # Projects
        some_day = datetime(year=self.year, month=2, day=27, tzinfo=pytz.UTC)
        project1 = ProjectFactory.create(
            owner=self.users[0],
            status=done_complete,
            campaign_ended=some_day
        )
        ProjectFactory.create(owner=self.users[0], status=done_incomplete)

        # Tasks
        task = TaskFactory.create(author=self.users[0], project=project1, people_needed=2, status='realized')

        for month in range(1, 12):
            for x in range(1, 10):

                task_member = TaskMemberFactory.create(
                    time_spent=10,
                    member=self.users[month * 10 + x],
                    task=task,
                    status=TaskMember.TaskMemberStatuses.applied
                )
                task_member.status = TaskMember.TaskMemberStatuses.realized
                task_member.save()

        # xls export
        self.xls_file_name = 'test.xlsx'
        self.xls_file_path = os.path.join(settings.PROJECT_ROOT, self.xls_file_name)
        self.command = ParticipationCommand()

    def test_export(self):
        with patch.object(self.command, 'get_xls_file_name', return_value=self.xls_file_name):
            call_command(self.command, '--start', self.year, '--end', self.year, '--tenant', 'test')
            self.assertTrue(os.path.isfile(self.xls_file_path), True)
            workbook = load_workbook(filename=self.xls_file_path, read_only=True)

            # Check participants
            self.assertEqual(workbook.worksheets[0]['A1'].value, 'Email Address')
            self.assertEqual(workbook.worksheets[0]['A2'].value, self.users[0].email)
            self.assertEqual(workbook.worksheets[0]['A3'].value, self.users[11].email)

            # Check some sheet titles
            self.assertEqual(workbook.worksheets[0].title, 'Participants - {}'.format(self.year))
            self.assertEqual(workbook.worksheets[1].title, 'Totals - {}'.format(self.year))
            self.assertEqual(workbook.worksheets[6].title, 'Location Segmentation - {}'.format(self.year))
            self.assertEqual(workbook.worksheets[7].title, 'Theme Segmentation - {}'.format(self.year))


class TestExportAnalytics(BluebottleTestCase):

    def setUp(self):
        super(TestExportAnalytics, self).setUp()
        self.init_projects()
        self.year = datetime.now().year

        # Project Phases
        done_complete = ProjectPhase.objects.get(slug="done-complete")
        done_incomplete = ProjectPhase.objects.get(slug="done-incomplete")

        # Users
        self.users = BlueBottleUserFactory.create_batch(200)

        # Projects
        some_day = datetime(year=self.year, month=2, day=27, tzinfo=pytz.UTC)
        project1 = ProjectFactory.create(
            owner=self.users[0],
            status=done_complete,
            campaign_ended=some_day
        )
        ProjectFactory.create(owner=self.users[0], status=done_incomplete)

        # Tasks
        task = TaskFactory.create(author=self.users[0], project=project1, people_needed=2, status='realized')

        for month in range(1, 12):
            for x in range(1, 10):

                task_member = TaskMemberFactory.create(
                    time_spent=10,
                    member=self.users[month * 10 + x],
                    task=task,
                    status=TaskMember.TaskMemberStatuses.applied
                )
                task_member.status = TaskMember.TaskMemberStatuses.realized
                task_member.save()

        # xls export
        self.xls_file_name = 'test.xlsx'
        self.xls_file_path = os.path.join(settings.PROJECT_ROOT, self.xls_file_name)
        self.command = AnalyticsCommand()

    @patch('bluebottle.analytics.utils.queue_analytics_record', return_value=True)
    def test_export(self, mock_queue):
        call_command(self.command,
                     '--start', '%s-01-01' % self.year,
                     '--end', '%s-12-31' % self.year,
                     '--tenant', 'test')
        self.assertEqual(mock_queue.call_count, 403)
