import json
import urllib
from datetime import timedelta, date
from io import BytesIO

import icalendar
from django.contrib.auth.models import Group, Permission
from django.contrib.gis.geos import Point
from django.urls import reverse
from django.utils.timezone import now
from openpyxl import load_workbook
from rest_framework import status
from pytz import UTC

from bluebottle.files.tests.factories import PrivateDocumentFactory
from bluebottle.initiatives.models import InitiativePlatformSettings
from bluebottle.initiatives.tests.factories import InitiativeFactory, InitiativePlatformSettingsFactory
from bluebottle.members.models import MemberPlatformSettings
from bluebottle.segments.tests.factories import SegmentTypeFactory, SegmentFactory
from bluebottle.test.factory_models.accounts import BlueBottleUserFactory
from bluebottle.test.factory_models.geo import LocationFactory, PlaceFactory, GeolocationFactory
from bluebottle.test.factory_models.projects import ThemeFactory
from bluebottle.test.utils import (
    APITestCase
)
from bluebottle.test.utils import BluebottleTestCase, JSONAPITestClient, get_first_included_by_type
from bluebottle.time_based.models import SlotParticipant, Skill
from bluebottle.time_based.tests.factories import (
    DateActivityFactory,
    DateParticipantFactory,
    DateActivitySlotFactory, SlotParticipantFactory, SkillFactory
)


class TimeBasedListAPIViewTestCase():
    def setUp(self):
        super().setUp()
        self.settings = InitiativePlatformSettingsFactory.create(
            activity_types=[self.factory._meta.model.__name__.lower()]
        )

        self.client = JSONAPITestClient()
        self.url = reverse('{}-list'.format(self.type))
        self.user = BlueBottleUserFactory()
        self.initiative = InitiativeFactory(owner=self.user, status='approved')

        self.data = {
            'data': {
                'type': 'activities/time-based/{}s'.format(self.type),
                'attributes': {
                    'title': 'Beach clean-up Katwijk',
                    'review': False,
                    'is-online': True,
                    'registration-deadline': str(date.today() + timedelta(days=14)),
                    'capacity': 10,
                    'description': 'We will clean up the beach south of Katwijk'
                },
                'relationships': {
                    'initiative': {
                        'data': {
                            'type': 'initiatives', 'id': self.initiative.id
                        },
                    },
                }
            }
        }

    def test_create_complete(self):
        response = self.client.post(self.url, json.dumps(self.data), user=self.user)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        self.response_data = response.json()['data']

        self.assertEqual(self.response_data['attributes']['status'], 'draft')
        self.assertEqual(self.response_data['attributes']['title'], self.data['data']['attributes']['title'])
        self.assertEqual(
            self.response_data['meta']['permissions']['GET'],
            True
        )

        self.assertEqual(
            self.response_data['meta']['permissions']['PUT'],
            True
        )

        self.assertEqual(
            self.response_data['meta']['permissions']['PATCH'],
            True
        )

    def test_create_duplicate_title(self):
        DateActivityFactory.create(
            title=self.data['data']['attributes']['title']
        )

        # Add an activity with the same title should NOT return an error
        response = self.client.post(self.url, json.dumps(self.data), user=self.user)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_create_disabled(self):
        self.settings.activity_types = ('funding',)
        self.settings.save()

        response = self.client.post(self.url, json.dumps(self.data), user=self.user)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_create_no_title(self):
        del self.data['data']['attributes']['title']
        response = self.client.post(self.url, json.dumps(self.data), user=self.user)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        self.assertTrue(
            '/data/attributes/title' in (
                error['source']['pointer'] for error in response.json()['data']['meta']['required']
            )
        )

    def test_create_as_activity_manager(self):
        activity_manager = BlueBottleUserFactory.create()
        self.initiative.activity_managers.add(activity_manager)

        response = self.client.post(self.url, json.dumps(self.data), user=activity_manager)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_create_not_initiator(self):
        another_user = BlueBottleUserFactory.create()
        response = self.client.post(self.url, json.dumps(self.data), user=another_user)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_create_not_initiator_open(self):
        self.initiative.is_open = True
        self.initiative.save()
        another_user = BlueBottleUserFactory.create()
        response = self.client.post(self.url, json.dumps(self.data), user=another_user)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_create_not_initiator_not_approved(self):
        self.initiative.is_open = True
        self.initiative.status = 'draft'
        self.initiative.save()

        another_user = BlueBottleUserFactory.create()
        response = self.client.post(self.url, json.dumps(self.data), user=another_user)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


class DateListAPIViewTestCase(TimeBasedListAPIViewTestCase, BluebottleTestCase):
    type = 'date'
    factory = DateActivityFactory
    participant_factory = DateParticipantFactory

    def setUp(self):
        super().setUp()
        self.slot_url = reverse('date-slot-list')
        self.data['data']['attributes'].update({
            'start': str(now() + timedelta(days=21)),
            'duration': '4:00:00',
        })

        self.slot_data = {
            'data': {
                'type': 'activities/time-based/date-slots',
                'attributes': {
                    'title': 'Kick-off',
                    'is-online': True,
                    'start': '2020-12-01T10:00:00+01:00',
                    'duration': '2:30:00',
                    'capacity': 10,
                },
                'relationships': {
                    'activity': {
                        'data': {
                            'type': 'activities/time-based/dates',
                            'id': 0
                        },
                    },
                }
            }
        }

    def test_create_complete(self):
        super().test_create_complete()
        # Can't yet submit because we don't have a slot yet
        self.assertEqual(
            {
                transition['name'] for transition in
                self.response_data['meta']['transitions']
            },
            {'delete'}
        )

    def test_add_slots_by_owner(self):
        response = self.client.post(self.url, json.dumps(self.data), user=self.user)
        self.response_data = response.json()['data']
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        activity_id = response.json()['data']['id']
        self.slot_data['data']['relationships']['activity']['data']['id'] = activity_id
        response = self.client.post(self.slot_url, json.dumps(self.slot_data), user=self.user)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.slot_data['data']['attributes']['title'] = 'Second meeting'
        self.slot_data['data']['attributes']['start'] = '2020-12-05T10:00:00+01:00'
        response = self.client.post(self.slot_url, json.dumps(self.slot_data), user=self.user)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        activity_url = reverse('date-detail', args=(activity_id,))
        response = self.client.get(activity_url, user=self.user)
        self.response_data = response.json()['data']
        # Now we can publish the activity
        self.assertEqual(
            {
                transition['name'] for transition in
                self.response_data['meta']['transitions']
            },
            {'publish', 'delete'}
        )

    def test_add_slots_by_other(self):
        response = self.client.post(self.url, json.dumps(self.data), user=self.user)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        activity_id = response.json()['data']['id']
        self.slot_data['data']['relationships']['activity']['data']['id'] = activity_id
        other = BlueBottleUserFactory.create()
        response = self.client.post(self.slot_url, json.dumps(self.slot_data), user=other)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


class TimeBasedDetailAPIViewTestCase():
    def setUp(self):
        super().setUp()
        self.settings = InitiativePlatformSettings.load()
        self.settings.activity_types = [self.factory._meta.model.__name__.lower()]
        self.settings.save()

        self.client = JSONAPITestClient()
        self.user = BlueBottleUserFactory()
        self.activity = self.factory.create()
        self.activity.refresh_from_db()

        self.url = reverse('{}-detail'.format(self.type), args=(self.activity.pk,))

        self.data = {
            'data': {
                'type': 'activities/time-based/{}s'.format(self.type),
                'id': str(self.activity.pk),
                'attributes': {
                    'title': 'Beach clean-up Katwijk',
                    'review': False,
                    'is-online': True,
                    'registration-deadline': str(date.today() + timedelta(days=14)),
                    'capacity': 10,
                    'description': 'We will clean up the beach south of Katwijk'
                },
                'relationships': {
                    'initiative': {
                        'data': {
                            'type': 'initiatives', 'id': self.activity.initiative.id
                        },
                    },
                }
            }
        }

    def assertTransitionInData(self, transition, data):
        self.assertIn(transition, [trans['name'] for trans in data['meta']['transitions']])

    def test_get_owner(self):
        self.activity.initiative.states.submit(save=True)
        self.activity.initiative.states.approve(save=True)

        response = self.client.get(self.url, user=self.activity.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()['data']
        self.assertEqual(data['attributes']['title'], self.activity.title)

        self.assertEqual(
            data['meta']['permissions']['GET'],
            True
        )

        self.assertEqual(
            data['meta']['permissions']['PUT'],
            True
        )

        self.assertEqual(
            data['meta']['permissions']['PATCH'],
            True
        )
        self.assertTransitionInData('cancel', data)
        self.assertEqual(data['meta']['matching-properties']['skill'], None)
        self.assertEqual(data['meta']['matching-properties']['theme'], None)
        self.assertEqual(data['meta']['matching-properties']['location'], None)

        contributor_url = reverse(f'{self.type}-participants', args=(self.activity.pk,))

        self.assertTrue(
            f"{contributor_url}?filter[status]=new" in
            data['relationships']['unreviewed-contributors']['links']['related']
        )
        self.assertTrue(
            contributor_url in data['relationships']['contributors']['links']['related']
        )

    def test_matching_theme(self):
        self.activity.initiative.states.submit(save=True)
        self.activity.initiative.states.approve(save=True)

        user = BlueBottleUserFactory.create()
        user.favourite_themes.add(self.activity.initiative.theme)

        response = self.client.get(self.url, user=user)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()['data']

        self.assertEqual(data['meta']['matching-properties']['skill'], None)
        self.assertEqual(data['meta']['matching-properties']['theme'], True)
        self.assertEqual(data['meta']['matching-properties']['location'], None)

    def test_mismatching_theme(self):
        self.activity.initiative.states.submit(save=True)
        self.activity.initiative.states.approve(save=True)

        user = BlueBottleUserFactory.create()
        user.favourite_themes.add(ThemeFactory.create())

        response = self.client.get(self.url, user=user)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()['data']

        self.assertEqual(data['meta']['matching-properties']['skill'], None)
        self.assertEqual(data['meta']['matching-properties']['theme'], False)
        self.assertEqual(data['meta']['matching-properties']['location'], None)

    def test_matching_skill(self):
        self.activity.initiative.states.submit(save=True)
        self.activity.initiative.states.approve(save=True)

        user = BlueBottleUserFactory.create()
        user.skills.add(self.activity.expertise)

        response = self.client.get(self.url, user=user)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()['data']

        self.assertEqual(data['meta']['matching-properties']['skill'], True)
        self.assertEqual(data['meta']['matching-properties']['theme'], None)
        self.assertEqual(data['meta']['matching-properties']['location'], None)

    def test_mismatching_skill(self):
        self.activity.initiative.states.submit(save=True)
        self.activity.initiative.states.approve(save=True)

        user = BlueBottleUserFactory.create()
        user.skills.add(SkillFactory.create())

        response = self.client.get(self.url, user=user)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()['data']

        self.assertEqual(data['meta']['matching-properties']['skill'], False)
        self.assertEqual(data['meta']['matching-properties']['theme'], None)
        self.assertEqual(data['meta']['matching-properties']['location'], None)

    def test_get_owner_export_disabled(self):
        initiative_settings = InitiativePlatformSettings.load()
        initiative_settings.enable_participant_exports = False
        initiative_settings.save()
        response = self.client.get(self.url, user=self.activity.owner)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()['data']
        export_url = data['attributes']['participants-export-url']
        self.assertIsNone(export_url)

    def test_get_owner_export_enabled(self):
        self.participant_factory.create_batch(4, activity=self.activity)

        initiative_settings = InitiativePlatformSettings.load()
        initiative_settings.enable_participant_exports = True
        initiative_settings.save()
        self.activity.review_title = 'Motivation'
        self.activity.save()

        response = self.client.get(self.url, user=self.activity.owner)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()['data']
        export_url = data['attributes']['participants-export-url']['url']
        export_response = self.client.get(export_url)

        workbook = load_workbook(filename=BytesIO(export_response.content))
        self.assertEqual(len(workbook.worksheets), 1)

        sheet = workbook.get_active_sheet()

        slot = self.activity.slots.first()
        self.assertEqual(
            tuple(sheet.values)[0],
            ('Email', 'Name', 'Registration Date', 'Status', 'Motivation')
        )
        self.assertEqual(
            sheet.title,
            f'{slot.start.strftime("%d-%m-%y %H%M")} {slot.id} {slot.title}'[:29]
        )

        wrong_signature_response = self.client.get(export_url + '111')
        self.assertEqual(
            wrong_signature_response.status_code, 404
        )

    def test_export_invalid_charachter(self):
        self.activity.title = 'test/* - *)'
        self.activity.save()

        initiative_settings = InitiativePlatformSettings.load()
        initiative_settings.enable_participant_exports = True
        initiative_settings.save()

        response = self.client.get(self.url, user=self.activity.owner)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()['data']
        export_url = data['attributes']['participants-export-url']['url']
        export_response = self.client.get(export_url)

        self.assertEqual(export_response.status_code, status.HTTP_200_OK)

    def test_export_with_segments(self):
        initiative_settings = InitiativePlatformSettings.load()
        initiative_settings.enable_participant_exports = True
        initiative_settings.save()

        department = SegmentTypeFactory.create(name='Department')
        music = SegmentTypeFactory.create(name='Music')
        workshop = SegmentFactory.create(
            segment_type=department,
            name='Workshop'
        )
        metal = SegmentFactory.create(
            segment_type=music,
            name='Metal'
        )
        classical = SegmentFactory.create(
            segment_type=music,
            name='Classical'
        )
        user = BlueBottleUserFactory.create()
        user.segments.add(workshop)
        user.segments.add(metal)
        user.segments.add(classical)
        participant = self.participant_factory.create(
            activity=self.activity,
            user=user,
            status='accepted'
        )
        for slot in self.activity.slots.all():
            SlotParticipantFactory.create(
                slot=slot,
                participant=participant,
                status='registered'
            )
        self.activity.review_title = "Motivation"
        self.activity.save()

        response = self.client.get(self.url, user=self.activity.owner)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()['data']
        export_url = data['attributes']['participants-export-url']['url']
        export_response = self.client.get(export_url)
        sheet = load_workbook(
            filename=BytesIO(export_response.content)
        ).get_active_sheet()
        self.assertEqual(sheet["A1"].value, "Email")
        self.assertEqual(sheet["B1"].value, "Name")
        self.assertEqual(sheet["E1"].value, "Motivation")
        self.assertEqual(sheet["F1"].value, "Department")
        self.assertEqual(sheet["G1"].value, "Music")

        self.assertEqual(sheet['F2'].value, 'Workshop')
        self.assertEqual(sheet['G2'].value, 'Classical, Metal')

    def test_get_other_user_export(self):
        response = self.client.get(self.url, user=self.user)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()['data']
        export_url = data['attributes']['participants-export-url']
        self.assertIsNone(export_url)

    def test_get_open(self):
        self.activity.initiative.states.submit(save=True)
        self.activity.initiative.states.approve(save=True)
        if self.activity.states.submit:
            self.activity.states.publish(save=True)
        else:
            self.activity.states.publish(save=True)

        response = self.client.get(self.url, user=self.activity.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.data = response.json()['data']
        self.assertTransitionInData('cancel', self.data)

    def test_get_contributors(self):
        participants = self.participant_factory.create_batch(4, activity=self.activity)
        withdrawn = self.participant_factory.create(activity=self.activity)
        withdrawn.states.withdraw(save=True)

        for participant in [withdrawn] + participants:
            SlotParticipantFactory.create(
                participant=participant, slot=self.activity.slots.get()
            )

        response = self.client.get(self.url, user=self.activity.owner)

        self.response_data = response.json()['data']
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.assertEqual(
            self.response_data['meta']['contributor-count'],
            4
        )
        response = self.client.get(
            self.response_data['relationships']['contributors']['links']['related'],
            user=self.activity.owner
        )
        self.response_data = response.json()['data']
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.assertEqual(
            len(self.response_data),
            5
        )

    def test_get_contributors_anonymous(self):
        participants = self.participant_factory.create_batch(4, activity=self.activity)
        withdrawn = self.participant_factory.create(activity=self.activity)
        withdrawn.states.withdraw(save=True)

        for participant in [withdrawn] + participants:
            SlotParticipantFactory.create(
                participant=participant, slot=self.activity.slots.get()
            )

        response = self.client.get(self.url)

        self.response_data = response.json()['data']
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.assertEqual(
            self.response_data['meta']['contributor-count'],
            4
        )

        response = self.client.get(
            self.response_data['relationships']['contributors']['links']['related'],
        )
        self.response_data = response.json()['data']
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.assertEqual(
            len(self.response_data),
            4
        )

    def test_get_contributors_participant(self):
        participants = self.participant_factory.create_batch(4, activity=self.activity)
        withdrawn = self.participant_factory.create(activity=self.activity)
        withdrawn.states.withdraw(save=True)
        used_participant = self.participant_factory.create(activity=self.activity)
        used_participant.states.withdraw(save=True)

        for participant in [withdrawn, used_participant] + participants:
            SlotParticipantFactory.create(
                participant=participant, slot=self.activity.slots.get()
            )

        response = self.client.get(self.url, user=used_participant.user)

        self.response_data = response.json()['data']
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.assertEqual(
            self.response_data['meta']['contributor-count'],
            4
        )

        response = self.client.get(
            self.response_data["relationships"]["contributors"]["links"]["related"],
            user=used_participant.user,
        )
        self.response_data = response.json()['data']
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.assertEqual(
            len(self.response_data),
            5
        )

    def test_get_non_anonymous(self):
        response = self.client.get(self.url)

        data = response.json()['data']
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(data['attributes']['title'], self.activity.title)

        self.assertEqual(
            data['meta']['permissions']['GET'],
            True
        )

        self.assertEqual(
            data['meta']['permissions']['PUT'],
            False
        )

        self.assertEqual(
            data['meta']['permissions']['PATCH'],
            False
        )

    def test_get_my_contributor(self):
        participant = self.participant_factory.create(activity=self.activity)
        response = self.client.get(self.url, user=participant.user)

        included_participant = get_first_included_by_type(
            response, self.participant_factory._meta.model.JSONAPIMeta.resource_name
        )
        self.assertEqual(str(participant.pk), included_participant['id'])

    def test_update_owner(self):
        response = self.client.put(self.url, json.dumps(self.data), user=self.activity.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response.json()['data']['attributes']['title'],
            self.data['data']['attributes']['title']
        )

    def test_update_manager(self):
        response = self.client.put(
            self.url, json.dumps(self.data), user=self.activity.initiative.activity_managers.first()
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response.json()['data']['attributes']['title'],
            self.data['data']['attributes']['title']
        )

    def test_update_initiative_owner(self):
        response = self.client.put(
            self.url, json.dumps(self.data), user=self.activity.initiative.owner
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response.json()['data']['attributes']['title'],
            self.data['data']['attributes']['title']
        )

    def test_update_unauthenticated(self):
        response = self.client.put(self.url, json.dumps(self.data))

        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_update_wrong_user(self):
        response = self.client.put(
            self.url, json.dumps(self.data), user=BlueBottleUserFactory.create()
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_delete_owner(self):
        response = self.client.delete(self.url, user=self.activity.owner)

        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)

    def test_delete_unauthenticated(self):
        response = self.client.delete(self.url)

        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_delete_wrong_user(self):
        response = self.client.delete(
            self.url, user=BlueBottleUserFactory.create()
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_delete_submitted(self):
        self.activity.initiative.states.submit(save=True)
        response = self.client.delete(
            self.url, user=self.activity.owner
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_update_deleted(self):
        self.activity.states.delete(save=True)
        response = self.client.put(self.url, json.dumps(self.data), user=self.activity.owner)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_update_rejected(self):
        self.activity.states.reject(save=True)
        response = self.client.put(self.url, json.dumps(self.data), user=self.activity.owner)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


class DateDetailAPIViewTestCase(TimeBasedDetailAPIViewTestCase, BluebottleTestCase):
    type = 'date'
    factory = DateActivityFactory
    participant_factory = DateParticipantFactory

    def setUp(self):
        super().setUp()

        self.data['data']['attributes'].update({
            'start': str(now() + timedelta(days=21)),
            'duration': '4:00',
        })
        self.slot = self.activity.slots.first()
        self.slot_url = reverse('date-slot-detail', args=(self.slot.pk,))

    def test_get_calendar_links(self):
        response = self.client.get(self.url, user=self.activity.owner)

        links = response.json()['data']['attributes']['links']

        self.assertTrue(
            links['ical'].startswith(
                reverse('date-ical', args=(self.activity.pk, self.activity.owner.id))
            )
        )

    def test_matching_all(self):
        self.activity.initiative.states.submit(save=True)
        self.activity.initiative.states.approve(save=True)

        location = GeolocationFactory.create(
            position=Point(x=4.8981734, y=52.3790565)
        )
        slot = self.activity.slots.first()
        slot.location = location
        slot.save()

        user = BlueBottleUserFactory.create()

        user.place = PlaceFactory.create(
            position=Point(x=4.8981730, y=52.3790560)
        )
        user.skills.add(self.activity.expertise)
        user.favourite_themes.add(self.activity.initiative.theme)
        user.save()

        response = self.client.get(self.url, user=user)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()['data']

        self.assertEqual(data['meta']['matching-properties']['skill'], True)
        self.assertEqual(data['meta']['matching-properties']['theme'], True)
        self.assertEqual(data['meta']['matching-properties']['location'], True)

    def test_matching_all_cancelled(self):
        self.activity.initiative.states.submit(save=True)
        self.activity.initiative.states.approve(save=True)

        self.activity.refresh_from_db()
        self.activity.states.cancel(save=True)

        location = GeolocationFactory.create(
            position=Point(x=4.8981734, y=52.3790565)
        )
        slot = self.activity.slots.first()
        slot.location = location
        slot.save()

        user = BlueBottleUserFactory.create()

        PlaceFactory.create(
            content_object=user,
            position=Point(x=4.9848386, y=52.3929661)
        )

        user.skills.add(self.activity.expertise)
        user.favourite_themes.add(self.activity.initiative.theme)

        response = self.client.get(self.url, user=user)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()['data']

        self.assertEqual(data['meta']['matching-properties']['skill'], False)
        self.assertEqual(data['meta']['matching-properties']['theme'], False)
        self.assertEqual(data['meta']['matching-properties']['location'], False)

    def test_matching_location_place(self):
        self.activity.initiative.states.submit(save=True)
        self.activity.initiative.states.approve(save=True)

        location = GeolocationFactory.create(
            position=Point(x=4.8981734, y=52.3790565)
        )
        slot = self.activity.slots.first()
        slot.location = location
        slot.save()

        user = BlueBottleUserFactory.create()
        user.place = PlaceFactory.create(
            position=Point(x=4.9848386, y=52.3929661)
        )
        user.save()

        response = self.client.get(self.url, user=user)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()['data']

        self.assertEqual(data['meta']['matching-properties']['skill'], None)
        self.assertEqual(data['meta']['matching-properties']['theme'], None)
        self.assertEqual(data['meta']['matching-properties']['location'], True)

    def test_matching_location_location(self):
        self.activity.initiative.states.submit(save=True)
        self.activity.initiative.states.approve(save=True)

        location = GeolocationFactory.create(
            position=Point(
                x=4.8981734, y=52.3790565
            )
        )
        slot = self.activity.slots.first()
        slot.location = location
        slot.save()

        user = BlueBottleUserFactory.create(
            location=LocationFactory.create(
                position=Point(x=4.9848386, y=52.3929661)
            )
        )
        response = self.client.get(self.url, user=user)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()['data']

        self.assertEqual(data['meta']['matching-properties']['skill'], None)
        self.assertEqual(data['meta']['matching-properties']['theme'], None)
        self.assertEqual(data['meta']['matching-properties']['location'], True)

    def test_matching_location_place_too_far(self):
        self.activity.initiative.states.submit(save=True)
        self.activity.initiative.states.approve(save=True)

        location = GeolocationFactory.create(
            position=Point(x=4.4207882, y=51.9280712)
        )
        slot = self.activity.slots.first()
        slot.location = location
        slot.save()

        user = BlueBottleUserFactory.create()
        user.place = PlaceFactory.create(
            position=Point(x=4.9848386, y=52.3929661)
        )
        user.save()

        response = self.client.get(self.url, user=user)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()['data']

        self.assertEqual(data['meta']['matching-properties']['skill'], None)
        self.assertEqual(data['meta']['matching-properties']['theme'], None)
        self.assertEqual(data['meta']['matching-properties']['location'], False)

    def test_matching_location_location_too_far(self):
        self.activity.initiative.states.submit(save=True)
        self.activity.initiative.states.approve(save=True)

        location = GeolocationFactory(
            position=Point(x=4.4207882, y=51.9280712)
        )
        slot = self.activity.slots.first()
        slot.location = location
        slot.save()

        user = BlueBottleUserFactory.create(
            location=LocationFactory.create(
                position=Point(x=4.9848386, y=52.3929661)
            )
        )

        response = self.client.get(self.url, user=user)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()['data']

        self.assertEqual(data['meta']['matching-properties']['skill'], None)
        self.assertEqual(data['meta']['matching-properties']['theme'], None)
        self.assertEqual(data['meta']['matching-properties']['location'], False)


class TimeBasedTransitionAPIViewTestCase():
    def setUp(self):
        super().setUp()
        self.client = JSONAPITestClient()
        self.user = BlueBottleUserFactory()
        self.activity = self.factory.create()

        self.url = reverse('{}-transition-list'.format(self.type))
        self.data = {
            'data': {
                'type': 'activities/time-based/{}-transitions'.format(self.type),
                'attributes': {},
                'relationships': {
                    'resource': {
                        'data': {
                            'type': 'activities/time-based/{}s'.format(self.type),
                            'id': self.activity.pk
                        }
                    }
                }
            }
        }

    def test_delete_by_owner(self):
        # Owner can delete the event
        self.data['data']['attributes']['transition'] = 'delete'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.activity.owner
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        data = json.loads(response.content)

        self.assertEqual(
            data['included'][0]['type'],
            'activities/time-based/{}s'.format(self.type)
        )
        self.assertEqual(data['included'][0]['attributes']['status'], 'deleted')

    def test_delete_by_other_user(self):
        self.data['data']['attributes']['transition'] = 'delete'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=BlueBottleUserFactory.create()
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        data = json.loads(response.content)
        self.assertEqual(data['errors'][0], "Transition is not available")

    def test_reject(self):
        self.data['data']['attributes']['transition'] = 'reject'
        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.activity.owner
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        data = json.loads(response.content)
        self.assertEqual(data['errors'][0], "Transition is not available")

    def test_approve(self):
        self.data['data']['attributes']['transition'] = 'approve'
        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.activity.owner
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        data = json.loads(response.content)
        self.assertEqual(data['errors'][0], "Transition is not available")


class DateTransitionAPIViewTestCase(TimeBasedTransitionAPIViewTestCase, BluebottleTestCase):
    type = 'date'
    factory = DateActivityFactory
    participant_factory = DateParticipantFactory


class DateActivitySlotListAPITestCase(BluebottleTestCase):
    def setUp(self):
        self.client = JSONAPITestClient()

        self.url = reverse('date-slot-list')
        self.activity = DateActivityFactory.create(slots=[])

        self.data = {
            'data': {
                'type': 'activities/time-based/date-slots',
                'attributes': {
                    'title': 'Kick-off',
                    'is-online': True,
                    'start': '2020-12-01T10:00:00+01:00',
                    'duration': '2:30:00',
                    'capacity': 10,
                },
                'relationships': {
                    'activity': {
                        'data': {
                            'type': 'activities/time-based/dates',
                            'id': str(self.activity.pk)
                        },
                    },
                }
            }
        }

    def test_get(self):
        DateActivitySlotFactory.create_batch(3, activity=self.activity)
        DateActivitySlotFactory.create_batch(3, activity=DateActivityFactory.create())

        response = self.client.get(self.url, {'activity': self.activity.id})
        self.assertEqual(response.json()['meta']['pagination']['count'], len(self.activity.slots.all()))
        self.assertEqual(response.json()['meta']['total'], len(self.activity.slots.all()))

        slot_ids = [str(slot.pk) for slot in self.activity.slots.all()]
        for slot in response.json()['data']:
            self.assertTrue(slot['id'] in slot_ids)

    def test_get_filtered_start(self):
        DateActivitySlotFactory.create(
            start=now() + timedelta(days=2),
            activity=self.activity
        )
        DateActivitySlotFactory.create(
            start=now() + timedelta(days=4),
            activity=self.activity
        )
        latest = DateActivitySlotFactory.create(
            start=now() + timedelta(days=6),
            activity=self.activity
        )

        response = self.client.get(
            self.url,
            {
                'activity': self.activity.id,
                'start': (now() + timedelta(days=5)).strftime('%Y-%m-%d')
            }
        )
        self.assertEqual(response.json()['meta']['pagination']['count'], 1)
        self.assertEqual(response.json()['meta']['total'], len(self.activity.slots.all()))
        self.assertEqual(response.json()['data'][0]['id'], str(latest.pk))

    def test_get_invalid_start(self):
        DateActivitySlotFactory.create_batch(3, activity=self.activity)
        DateActivitySlotFactory.create_batch(3, activity=DateActivityFactory.create())

        response = self.client.get(
            self.url, {'activity': self.activity.id, 'start': 'invalid'}
        )
        self.assertEqual(response.json()['meta']['pagination']['count'], len(self.activity.slots.all()))
        self.assertEqual(response.json()['meta']['total'], len(self.activity.slots.all()))

        slot_ids = [str(slot.pk) for slot in self.activity.slots.all()]
        for slot in response.json()['data']:
            self.assertTrue(slot['id'] in slot_ids)

    def test_get_filtered_end(self):
        first = DateActivitySlotFactory.create(
            start=now() + timedelta(days=2),
            activity=self.activity
        )
        DateActivitySlotFactory.create(
            start=now() + timedelta(days=4),
            activity=self.activity
        )
        DateActivitySlotFactory.create(
            start=now() + timedelta(days=6),
            activity=self.activity
        )

        response = self.client.get(
            self.url,
            {
                'activity': self.activity.id,
                'end': (now() + timedelta(days=3)).strftime('%Y-%m-%d')
            }
        )
        self.assertEqual(response.json()['meta']['pagination']['count'], 1)
        self.assertEqual(response.json()['meta']['total'], len(self.activity.slots.all()))
        self.assertEqual(response.json()['data'][0]['id'], str(first.pk))

    def test_get_invalid_end(self):
        DateActivitySlotFactory.create_batch(3, activity=self.activity)
        DateActivitySlotFactory.create_batch(3, activity=DateActivityFactory.create())

        response = self.client.get(
            self.url, {'activity': self.activity.id, 'end': 'invalid'}
        )
        self.assertEqual(response.json()['meta']['pagination']['count'], len(self.activity.slots.all()))
        self.assertEqual(response.json()['meta']['total'], len(self.activity.slots.all()))

        slot_ids = [str(slot.pk) for slot in self.activity.slots.all()]
        for slot in response.json()['data']:
            self.assertTrue(slot['id'] in slot_ids)

    def test_get_filtered_both(self):
        DateActivitySlotFactory.create(
            start=now() + timedelta(days=2),
            activity=self.activity
        )
        middle = DateActivitySlotFactory.create(
            start=now() + timedelta(days=4),
            activity=self.activity
        )
        DateActivitySlotFactory.create(
            start=now() + timedelta(days=6),
            activity=self.activity
        )

        response = self.client.get(
            self.url,
            {
                'activity': self.activity.id,
                'start': (now() + timedelta(days=3)).strftime('%Y-%m-%d'),
                'end': (now() + timedelta(days=5)).strftime('%Y-%m-%d')
            }
        )
        self.assertEqual(response.json()['meta']['pagination']['count'], 1)
        self.assertEqual(response.json()['meta']['total'], len(self.activity.slots.all()))
        self.assertEqual(response.json()['data'][0]['id'], str(middle.pk))

    def test_get_filtered_contributor_id(self):
        participant = DateParticipantFactory.create(activity=self.activity)

        slot = DateActivitySlotFactory.create(
            start=now() + timedelta(days=2),
            activity=self.activity
        )

        slot_participant = SlotParticipantFactory(slot=slot, participant=participant)
        slot_participant.states.withdraw(save=True)

        second = DateActivitySlotFactory.create(
            start=now() + timedelta(days=4),
            activity=self.activity
        )
        slot_participant = SlotParticipantFactory(slot=second, participant=participant)

        third = DateActivitySlotFactory.create(
            start=now() + timedelta(days=6),
            activity=self.activity
        )
        other_participant = DateParticipantFactory.create(activity=self.activity)
        slot_participant = SlotParticipantFactory(slot=third, participant=other_participant)

        response = self.client.get(
            self.url,
            {
                'activity': self.activity.id,
                'contributor': participant.id
            }
        )
        self.assertEqual(response.json()['meta']['pagination']['count'], 1)
        self.assertEqual(response.json()['meta']['total'], 1)
        self.assertEqual(response.json()['data'][0]['id'], str(second.pk))

    def test_get_many(self):
        DateActivitySlotFactory.create_batch(12, activity=self.activity)
        DateActivitySlotFactory.create_batch(3, activity=DateActivityFactory.create())

        response = self.client.get(self.url, {'activity': self.activity.id})
        self.assertEqual(response.json()['meta']['pagination']['count'], len(self.activity.slots.all()))
        self.assertEqual(len(response.json()['data']), 8)

        slot_ids = [str(slot.pk) for slot in self.activity.slots.all()]
        for slot in response.json()['data']:
            self.assertTrue(slot['id'] in slot_ids)

        response = self.client.get(self.url, {'activity': self.activity.id, 'page[number]': 2})
        self.assertEqual(response.json()['meta']['pagination']['count'], len(self.activity.slots.all()))
        self.assertEqual(len(response.json()['data']), 4)

        slot_ids = [str(slot.pk) for slot in self.activity.slots.all()]
        for slot in response.json()['data']:
            self.assertTrue(slot['id'] in slot_ids)

    def test_get_no_activity_id(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_get_invalid_activity_id(self):
        response = self.client.get(self.url, {'activity': 'some-thing-wrong'})
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_get_incorrect_activity_id(self):
        response = self.client.get(self.url, {'activity': 1034320})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.json()['meta']['pagination']['count'], 0)
        self.assertEqual(len(response.json()['data']), 0)

    def test_get_closed_site(self):
        MemberPlatformSettings.objects.update(closed=True)
        group = Group.objects.get(name='Anonymous')
        group.permissions.remove(Permission.objects.get(codename='api_read_dateactivity'))
        group.permissions.remove(Permission.objects.get(codename='api_read_dateactivity'))

        response = self.client.get(self.url, {'activity': self.activity.pk})
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_create_owner(self):
        response = self.client.post(self.url, json.dumps(self.data), user=self.activity.owner)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        data = response.json()

        included = [{'id': resource['id'], 'type': resource['type']} for resource in data['included']]

        for attr in ['start', 'duration', 'capacity']:
            self.assertTrue(attr in data['data']['attributes'])

        self.assertEqual(data['data']['meta']['status'], 'finished')

        self.assertTrue(
            {'id': str(self.activity.pk), 'type': 'activities/time-based/dates'} in included
        )

    def test_create_other(self):
        response = self.client.post(
            self.url, json.dumps(self.data), user=BlueBottleUserFactory.create()
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_create_anonymous(self):
        response = self.client.post(self.url, json.dumps(self.data))

        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_create_open_activity(self):
        DateActivitySlotFactory.create(activity=self.activity)
        self.activity.initiative.states.submit()
        self.activity.initiative.states.approve(save=True)
        self.activity.states.publish(save=True)

        response = self.client.post(self.url, json.dumps(self.data), user=self.activity.owner)

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)


class DateActivitySlotDetailAPITestCase(BluebottleTestCase):
    def setUp(self):
        self.client = JSONAPITestClient()

        self.activity = DateActivityFactory.create()
        self.slot = DateActivitySlotFactory.create(activity=self.activity)

        self.url = reverse('date-slot-detail', args=(self.slot.pk,))
        self.data = {
            'data': {
                'type': 'activities/time-based/date-slots',
                'id': str(self.slot.pk),
                'attributes': {
                    'title': 'New title',
                },
            }
        }

    def test_update_owner(self):
        response = self.client.patch(self.url, json.dumps(self.data), user=self.activity.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()

        for attr in ['start', 'duration', 'capacity']:
            self.assertTrue(attr in data['data']['attributes'])

        self.assertEqual(data['data']['meta']['status'], 'open')

    def test_update_other(self):
        response = self.client.patch(
            self.url, json.dumps(self.data), user=BlueBottleUserFactory.create()
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_update_anonymous(self):
        response = self.client.patch(self.url, json.dumps(self.data))

        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_update_open_activity(self):
        self.activity.initiative.states.submit()
        self.activity.initiative.states.approve(save=True)
        self.activity.states.publish(save=True)

        response = self.client.patch(self.url, json.dumps(self.data), user=self.activity.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_get_owner(self):
        response = self.client.get(self.url, user=self.activity.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()

        for attr in ['start', 'duration', 'capacity']:
            self.assertTrue(attr in data['data']['attributes'])

        self.assertEqual(data['data']['meta']['status'], 'open')

    def test_get_other(self):
        response = self.client.get(
            self.url, user=BlueBottleUserFactory.create()
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_get_calendar_links(self):
        self.slot.location = None
        self.slot.is_online = True
        self.slot.online_meeting_url = 'http://example.com'
        self.slot.save()

        self.activity.description = "Test<br>bla"
        self.activity.save()

        response = self.client.get(self.url, user=self.activity.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        links = response.json()['data']['attributes']['links']

        self.assertTrue(
            links['ical'].startswith(
                reverse('slot-ical', args=(self.slot.pk,))
            )
        )

        params = urllib.parse.parse_qs(urllib.parse.urlparse(links['google']).query)
        self.assertEqual(params['action'], ['TEMPLATE'])
        self.assertEqual(params['text'][0], self.activity.title)

        self.assertEqual(
            params['details'][0],
            f'Test  \nbla\n\n{self.activity.get_absolute_url()}\nJoin: {self.slot.online_meeting_url}'
        )

    def test_get_calendar_links_location(self):
        self.slot.is_online = False
        self.slot.save()

        response = self.client.get(self.url, user=self.activity.owner)

        links = response.json()['data']['attributes']['links']

        params = urllib.parse.parse_qs(urllib.parse.urlparse(links['google']).query)
        self.assertEqual(
            params['location'][0], self.slot.location.formatted_address
        )

    def test_get_calendar_links_location_hint(self):
        self.slot.is_online = False
        self.slot.location_hint = 'On the second floor'
        self.slot.save()

        response = self.client.get(self.url, user=self.activity.owner)

        links = response.json()['data']['attributes']['links']

        params = urllib.parse.parse_qs(urllib.parse.urlparse(links['google']).query)
        self.assertEqual(
            params['location'][0],
            f'{self.slot.location.formatted_address} ({self.slot.location_hint})'
        )

    def test_closed_site(self):
        MemberPlatformSettings.objects.update(closed=True)
        group = Group.objects.get(name='Anonymous')
        group.permissions.remove(Permission.objects.get(codename='api_read_dateactivity'))
        group.permissions.remove(Permission.objects.get(codename='api_read_dateactivity'))
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_get_anonymous(self):
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_get_open_activity(self):
        self.activity.initiative.states.submit()
        self.activity.initiative.states.approve(save=True)
        self.activity.states.publish(save=True)

        response = self.client.get(self.url, user=self.activity.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_delete_owner(self):
        response = self.client.delete(self.url, user=self.activity.owner)

        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)

    def test_delete_other(self):
        response = self.client.delete(
            self.url, user=BlueBottleUserFactory.create()
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_delete_anonymous(self):
        response = self.client.delete(self.url)

        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_delete_open_activity(self):
        self.activity.initiative.states.submit()
        self.activity.initiative.states.approve(save=True)
        self.activity.states.publish(save=True)
        response = self.client.delete(self.url, user=self.activity.owner)
        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)


class ParticipantListViewTestCase():
    def setUp(self):
        super().setUp()
        self.client = JSONAPITestClient()
        self.user = BlueBottleUserFactory()
        self.activity = self.factory.create()

        self.url = reverse(self.url_name)

        self.private_document_url = reverse('private-document-list')
        self.png_document_path = './bluebottle/files/tests/files/test-image.png'

        self.data = {
            'data': {
                'type': self.participant_type,
                'attributes': {
                    'motiviation': 'I am great',
                },
                'relationships': {
                    'activity': {
                        'data': {
                            'type': 'activities/time-based/{}s'.format(self.type),
                            'id': self.activity.pk
                        }
                    }
                }
            }
        }

    def test_create(self):
        self.response = self.client.post(self.url, json.dumps(self.data), user=self.user)
        self.assertEqual(self.response.status_code, status.HTTP_201_CREATED)

        data = self.response.json()['data']
        self.assertEqual(
            data['relationships']['user']['data']['id'],
            str(self.user.pk)
        )

        self.assertEqual(
            data['meta']['permissions']['GET'],
            True
        )

        self.assertEqual(
            data['meta']['permissions']['PUT'],
            True
        )

        self.assertEqual(
            data['meta']['permissions']['PATCH'],
            True
        )

    def test_create_with_document(self):
        with open(self.png_document_path, 'rb') as test_file:
            document_response = self.client.post(
                self.private_document_url,
                test_file.read(),
                content_type="image/png",
                HTTP_CONTENT_DISPOSITION='attachment; filename="test.rtf"',
                user=self.user
            )

            self.assertEqual(document_response.status_code, 201)
            document_data = json.loads(document_response.content)

        self.data['data']['relationships']['document'] = {
            'data': {
                'type': 'private-documents',
                'id': document_data['data']['id']
            }
        }

        response = self.client.post(self.url, json.dumps(self.data), user=self.user)

        self.assertEqual(response.status_code, 201)

        data = response.json()['data']
        self.assertEqual(
            data['relationships']['document']['data']['id'],
            document_data['data']['id']
        )
        private_doc = self.included_by_type(response, 'private-documents')[0]
        self.assertTrue(
            private_doc['attributes']['link'].startswith(
                '{}?signature='.format(reverse(self.document_url_name, args=(data['id'],)))
            )
        )

    def test_create_duplicate(self):
        self.client.post(self.url, json.dumps(self.data), user=self.user)
        response = self.client.post(self.url, json.dumps(self.data), user=self.user)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            response.json()['errors'][0]['detail'],
            'The fields activity, user must make a unique set.'
        )

    def test_create_anonymous(self):
        response = self.client.post(self.url, json.dumps(self.data))

        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_get_participants(self):
        self.test_create()
        self.response = self.client.get(self.url)
        self.assertEqual(self.response.status_code, status.HTTP_200_OK)


class DateParticipantListAPIViewTestCase(ParticipantListViewTestCase, BluebottleTestCase):
    type = 'date'
    factory = DateActivityFactory
    participant_factory = DateParticipantFactory

    document_url_name = 'date-participant-document'
    application_type = 'contributions/time-based/date-participants'
    url_name = 'date-participant-list'
    participant_type = 'contributors/time-based/date-participants'

    def test_create(self):
        super().test_create()
        types = [included['type'] for included in self.response.json()['included']]
        self.assertTrue('members' in types)

    def test_get_participants(self):
        super().test_get_participants()
        types = [included['type'] for included in self.response.json()['included']]
        self.assertTrue('activities/time-based/dates' in types)
        self.assertTrue('members' in types)


class ParticipantDetailViewTestCase():
    def setUp(self):
        super().setUp()
        self.client = JSONAPITestClient()
        self.user = BlueBottleUserFactory()
        self.activity = self.factory.create()
        self.participant = self.participant_factory(
            activity=self.activity,
            motivation='My motivation'
        )

        self.url = reverse(self.url_name, args=(self.participant.pk,))

        self.private_document_url = reverse('private-document-list')
        self.png_document_path = './bluebottle/files/tests/files/test-image.png'

        self.data = {
            'data': {
                'type': self.participant_type,
                'id': self.participant.pk,
                'attributes': {'motivation': 'Let\'s go!!!'},
            }
        }

    def assertTransitionInData(self, transition, data):
        self.assertIn(transition, [trans['name'] for trans in data['meta']['transitions']])

    def test_get_user(self):
        response = self.client.get(self.url, user=self.participant.user)

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()['data']

        self.assertEqual(
            data['attributes']['motivation'],
            self.participant.motivation
        )

        self.assertEqual(
            data['relationships']['user']['data']['id'],
            str(self.participant.user.pk)
        )

        self.assertEqual(
            data['meta']['permissions']['GET'],
            True
        )

        self.assertEqual(
            data['meta']['permissions']['PUT'],
            True
        )

        self.assertEqual(
            data['meta']['permissions']['PATCH'],
            True
        )
        self.assertTransitionInData('withdraw', data)

    def test_get_owner(self):
        response = self.client.get(self.url, user=self.activity.owner)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.data = response.json()['data']

        self.assertEqual(
            self.data['attributes']['motivation'],
            self.participant.motivation
        )
        self.assertTransitionInData('remove', self.data)

    def test_get_participant(self):
        response = self.client.get(self.url, user=self.participant.user)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.data = response.json()['data']

        self.assertEqual(
            self.data['attributes']['motivation'],
            self.participant.motivation
        )
        self.assertTransitionInData('withdraw', self.data)

    def test_get_activity_manager(self):
        response = self.client.get(self.url, user=self.activity.initiative.activity_managers.first())

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()['data']
        self.assertEqual(
            data['attributes']['motivation'],
            self.participant.motivation
        )

    def test_get_other_user(self):
        response = self.client.get(self.url, user=self.user)

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()['data']

        self.assertIsNone(
            data['attributes']['motivation']
        )

    def test_patch_user(self):
        response = self.client.patch(self.url, json.dumps(self.data), user=self.participant.user)

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()['data']

        self.assertEqual(
            data['attributes']['motivation'],
            self.data['data']['attributes']['motivation']
        )

    def test_patch_document(self):
        with open(self.png_document_path, 'rb') as test_file:
            document_response = self.client.post(
                self.private_document_url,
                test_file.read(),
                content_type="image/png",
                HTTP_CONTENT_DISPOSITION='attachment; filename="test.rtf"',
                user=self.user
            )

            self.assertEqual(document_response.status_code, 201)
            document_data = json.loads(document_response.content)

        self.data['data']['relationships'] = {
            'document': {
                'data': {
                    'type': 'private-documents',
                    'id': document_data['data']['id']
                }
            }
        }

        response = self.client.patch(self.url, json.dumps(self.data), user=self.participant.user)

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()['data']

        self.assertEqual(
            data['relationships']['document']['data']['id'],
            document_data['data']['id']
        )

    def test_patch_other_user(self):
        response = self.client.patch(self.url, json.dumps(self.data), user=self.user)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_patch_anonymous(self):
        response = self.client.patch(self.url, json.dumps(self.data))

        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)


class DateParticipantDetailAPIViewTestCase(ParticipantDetailViewTestCase, BluebottleTestCase):
    type = 'date'
    factory = DateActivityFactory
    participant_factory = DateParticipantFactory
    url_name = 'date-participant-detail'
    participant_type = 'contributors/time-based/date-participants'


class ParticipantTransitionAPIViewTestCase():
    def setUp(self):
        super().setUp()
        self.client = JSONAPITestClient()
        self.user = BlueBottleUserFactory()
        self.activity = self.factory.create()
        self.participant = self.participant_factory.create(
            activity=self.activity
        )

        self.url = reverse(self.url_name)
        self.data = {
            'data': {
                'type': '{}-transitions'.format(self.participant_type),
                'attributes': {},
                'relationships': {
                    'resource': {
                        'data': {
                            'type': '{}s'.format(self.participant_type),
                            'id': self.participant.pk
                        }
                    }
                }
            }
        }

    def test_withdraw_by_user(self):
        self.data['data']['attributes']['transition'] = 'withdraw'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.participant.user
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        data = json.loads(response.content)

        participant = [
            include for include in data['included'] if include['type'] == '{}s'.format(self.participant_type)
        ]
        self.assertEqual(len(participant), 1)
        self.assertEqual(participant[0]['attributes']['status'], 'withdrawn')

    def test_withdraw_by_other_user(self):
        self.data['data']['attributes']['transition'] = 'withdraw'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.user
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_remove_by_activity_owner(self):
        self.data['data']['attributes']['transition'] = 'remove'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.activity.owner
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        data = json.loads(response.content)

        participant = [
            include for include in data['included'] if include['type'] == '{}s'.format(self.participant_type)
        ]
        self.assertEqual(len(participant), 1)
        self.assertEqual(participant[0]['attributes']['status'], 'removed')

    def test_remove_by_user(self):
        self.data['data']['attributes']['transition'] = 'remove'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.participant.user
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)


class DateParticipantTransitionAPIViewTestCase(ParticipantTransitionAPIViewTestCase, BluebottleTestCase):
    type = 'date'
    url_name = 'date-participant-transition-list'
    participant_type = 'contributors/time-based/date-participant'
    factory = DateActivityFactory
    participant_factory = DateParticipantFactory


class ReviewParticipantTransitionAPIViewTestCase():
    def setUp(self):
        super().setUp()
        self.client = JSONAPITestClient()
        self.user = BlueBottleUserFactory()
        self.another_user = BlueBottleUserFactory()
        self.activity = self.factory.create(review=True)
        self.participant = self.participant_factory.create(
            activity=self.activity,
            user=self.user,
            as_user=self.user
        )

        self.url = reverse(self.url_name)
        self.data = {
            'data': {
                'type': '{}-transitions'.format(self.participant_type),
                'attributes': {},
                'relationships': {
                    'resource': {
                        'data': {
                            'type': '{}s'.format(self.participant_type),
                            'id': self.participant.pk
                        }
                    }
                }
            }
        }

    def test_withdraw_by_user(self):
        self.data['data']['attributes']['transition'] = 'withdraw'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.user
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        data = json.loads(response.content)

        participant = [
            include for include in data['included'] if include['type'] == '{}s'.format(self.participant_type)
        ]
        self.assertEqual(len(participant), 1)
        self.assertEqual(participant[0]['attributes']['status'], 'withdrawn')

    def test_withdraw_by_other_user(self):
        self.data['data']['attributes']['transition'] = 'withdraw'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.another_user
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_reject_by_activity_owner(self):
        self.data['data']['attributes']['transition'] = 'reject'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.activity.owner
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        data = json.loads(response.content)

        participant = [
            include for include in data['included'] if include['type'] == '{}s'.format(self.participant_type)
        ]
        self.assertEqual(len(participant), 1)
        self.assertEqual(participant[0]['attributes']['status'], 'rejected')

    def test_reject_by_user(self):
        self.data['data']['attributes']['transition'] = 'reject'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.participant.user
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)


class DateReviewParticipantTransitionAPIViewTestCase(
    ReviewParticipantTransitionAPIViewTestCase, BluebottleTestCase
):
    type = 'date'
    url_name = 'date-participant-transition-list'
    participant_type = 'contributors/time-based/date-participant'
    factory = DateActivityFactory
    participant_factory = DateParticipantFactory


class RelatedParticipantsAPIViewTestCase():
    def setUp(self):
        super().setUp()
        self.client = JSONAPITestClient()
        self.activity = self.factory.create()
        self.participants = []
        for i in range(10):
            self.participants.append(
                self.participant_factory.create(
                    activity=self.activity,
                    document=PrivateDocumentFactory.create()
                )
            )

        self.participants[0].states.remove(save=True)
        self.participants[1].states.remove(save=True)

        self.url = reverse(self.url_name, args=(self.activity.pk,))

    def assertTotal(self, total):
        return self.assertEqual(self.response.json()['meta']['pagination']['count'], total)

    def test_get_owner(self):
        self.response = self.client.get(self.url, user=self.activity.owner)

        self.assertEqual(self.response.status_code, status.HTTP_200_OK)
        self.assertTotal(10)
        included_documents = self.included_by_type(self.response, 'private-documents')
        self.assertEqual(len(included_documents), 8)

    def test_get_owner_disable_last_name(self):
        MemberPlatformSettings.objects.update_or_create(display_member_names='first_name')
        self.response = self.client.get(self.url, user=self.activity.owner)

        self.assertEqual(self.response.status_code, status.HTTP_200_OK)

        for member in self.included_by_type(self.response, 'members'):
            self.assertTrue(member['attributes']['last-name'])

    def test_get_owner_disable_last_name_staff(self):
        MemberPlatformSettings.objects.update_or_create(display_member_names='first_name')
        staff = BlueBottleUserFactory.create(is_staff=True)
        self.response = self.client.get(self.url, user=staff)

        self.assertEqual(self.response.status_code, status.HTTP_200_OK)

        for member in self.included_by_type(self.response, 'members'):
            self.assertTrue(member['attributes']['last-name'])

    def test_get_with_duplicate_files(self):
        file = PrivateDocumentFactory.create(owner=self.participants[2].user)
        self.participants[2].document = file
        self.participants[2].save()
        self.participants[3].document = file
        self.participants[3].save()
        self.participants[4].document = file
        self.participants[4].save()
        self.response = self.client.get(self.url, user=self.activity.owner)
        self.assertEqual(self.response.status_code, status.HTTP_200_OK)
        self.assertTotal(10)
        included_documents = self.included_by_type(self.response, 'private-documents')
        self.assertEqual(len(included_documents), 6)

    def test_get_anonymous(self):
        self.response = self.client.get(self.url)

        self.assertEqual(self.response.status_code, status.HTTP_200_OK)
        self.assertTotal(8)
        included_documents = self.included_by_type(self.response, 'private-documents')
        self.assertEqual(len(included_documents), 0)

    def test_get_anonymous_disable_last_name(self):
        MemberPlatformSettings.objects.update_or_create(display_member_names='first_name')
        self.response = self.client.get(self.url)

        self.assertEqual(self.response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(self.response.json()['data']), 8)

        for member in self.included_by_type(self.response, 'members'):
            self.assertIsNone(member['attributes']['last-name'])

    def test_get_anonymous_disable_last_name_staff(self):
        MemberPlatformSettings.objects.update_or_create(display_member_names='first_name')
        staff = BlueBottleUserFactory.create(is_staff=True)
        self.response = self.client.get(self.url, user=staff)

        self.assertEqual(self.response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(self.response.json()['data']), 8)

        for member in self.included_by_type(self.response, 'members'):
            self.assertTrue(member['attributes']['last-name'])

    def test_get_removed_participant(self):
        self.response = self.client.get(self.url, user=self.participants[0].user)
        self.assertEqual(self.response.status_code, status.HTTP_200_OK)
        self.assertTotal(9)

    def test_get_filter_new(self):
        participant = self.participants[1]

        participant.status = 'new'
        participant.save()
        self.response = self.client.get(
            self.url + '?filter[status]=new', user=self.activity.owner
        )
        self.assertEqual(self.response.status_code, status.HTTP_200_OK)

        self.assertTotal(1)
        self.assertEqual(self.response.json()['data'][0]['id'], str(participant.pk))

    def test_get_filter_new_other_user(self):
        participant = self.participants[1]

        participant.status = 'new'
        participant.save()
        self.response = self.client.get(
            self.url + '?filter[status]=new', user=BlueBottleUserFactory.create()
        )
        self.assertEqual(self.response.status_code, status.HTTP_200_OK)

        self.assertTotal(0)

    def test_get_closed_site(self):
        MemberPlatformSettings.objects.update(closed=True)
        group = Group.objects.get(name='Anonymous')
        group.permissions.remove(Permission.objects.get(codename='api_read_dateparticipant'))
        group.permissions.remove(Permission.objects.get(codename='api_read_periodparticipant'))

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)


class RelatedDateParticipantAPIViewTestCase(RelatedParticipantsAPIViewTestCase, BluebottleTestCase):
    type = 'date'
    url_name = 'date-participants'
    participant_type = 'contributors/time-based/date-participant'
    factory = DateActivityFactory
    participant_factory = DateParticipantFactory

    def setUp(self):
        super().setUp()

        self.client = JSONAPITestClient()
        self.activity = self.factory.create()
        DateActivitySlotFactory.create(activity=self.activity)

        self.participants = []
        for i in range(10):
            participant = self.participant_factory.create(
                activity=self.activity,
                document=PrivateDocumentFactory.create()
            )
            for slot in self.activity.slots.all():
                SlotParticipantFactory.create(
                    participant=participant,
                    slot=slot
                )

            self.participants.append(
                participant
            )

        self.participants[0].states.remove(save=True)
        self.participants[1].states.remove(save=True)
        self.participants[2].slot_participants.all()[0].states.remove(save=True)

        self.url = reverse(self.url_name, args=(self.activity.pk,))

    def test_get_owner(self):
        super().test_get_owner()
        self.assertTotal(10)
        self.assertEqual(self.response.data['results'][0]['permissions']['PUT'], True)

    def test_get_anonymous(self):
        super().test_get_anonymous()
        self.assertTotal(8)
        self.assertEqual(self.response.data['results'][0]['permissions']['PUT'], False)

    def test_get_removed_participant(self):
        super().test_get_removed_participant()
        self.assertTotal(9)


class SlotParticipantListAPIViewTestCase(BluebottleTestCase):
    def setUp(self):
        super().setUp()
        self.client = JSONAPITestClient()
        self.activity = DateActivityFactory.create(review=False)
        self.slot = DateActivitySlotFactory.create(activity=self.activity)
        self.participant = DateParticipantFactory.create(activity=self.activity)

        self.url = reverse('slot-participant-list')

        self.data = {
            'data': {
                'attributes': {},
                'type': 'contributors/time-based/slot-participants',
                'relationships': {
                    'slot': {
                        'data': {
                            'type': 'activities/time-based/date-slots',
                            'id': self.slot.id
                        },
                    },
                    'participant': {
                        'data': {
                            'type': 'contributors/time-based/date-participants',
                            'id': self.participant.id
                        },
                    },
                }
            }
        }

    def test_create_participant_user(self):
        response = self.client.post(self.url, json.dumps(self.data), user=self.participant.user)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        data = response.json()['data']

        self.assertEqual(
            data['relationships']['slot']['data']['id'], str(self.slot.pk)
        )

        self.assertEqual(
            data['relationships']['participant']['data']['id'], str(self.participant.pk)
        )

        self.assertEqual(data['id'], str(self.participant.slot_participants.get().pk))

    def test_create_participant_user_full(self):
        self.slot.capacity = 1
        self.slot.save()

        part = DateParticipantFactory.create(activity=self.activity)
        SlotParticipantFactory.create(slot=self.slot, participant=part)
        response = self.client.post(self.url, json.dumps(self.data), user=self.participant.user)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_create_participant_user_twice(self):
        response = self.client.post(self.url, json.dumps(self.data), user=self.participant.user)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        response = self.client.post(self.url, json.dumps(self.data), user=self.participant.user)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_create_different_user(self):
        response = self.client.post(self.url, json.dumps(self.data), user=BlueBottleUserFactory.create())
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_create_activity_owner(self):
        response = self.client.post(self.url, json.dumps(self.data), user=self.activity.owner)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_create_no_user(self):
        response = self.client.post(self.url, json.dumps(self.data))
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_create_different_slot(self):
        activity = DateActivityFactory.create()
        slot = DateActivitySlotFactory.create(activity=activity)
        self.data['data']['relationships']['slot']['data']['id'] = slot.pk
        response = self.client.post(self.url, json.dumps(self.data), user=self.participant.user)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_create_missing_slot(self):
        del self.data['data']['relationships']['slot']

        response = self.client.post(self.url, json.dumps(self.data), user=self.participant.user)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_create_without_participant(self):
        # This will create a participant on the fly
        user = BlueBottleUserFactory.create()
        del self.data['data']['relationships']['participant']
        response = self.client.post(self.url, json.dumps(self.data), user=user)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_create_with_email_by_staff(self):
        # For bulk add flow
        BlueBottleUserFactory.create(email='rene@froger.nl')
        staff = BlueBottleUserFactory.create(is_staff=True)
        del self.data['data']['relationships']['participant']
        self.data['data']['attributes']['email'] = 'rene@froger.nl'
        response = self.client.post(self.url, json.dumps(self.data), user=staff)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_create_with_email_by_user(self):
        BlueBottleUserFactory.create(email='rene@froger.nl')
        user = BlueBottleUserFactory.create()
        del self.data['data']['relationships']['participant']
        self.data['data']['attributes']['email'] = 'rene@froger.nl'
        response = self.client.post(self.url, json.dumps(self.data), user=user)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


class SlotParticipantDetailAPIViewTestCase(BluebottleTestCase):
    def setUp(self):
        super().setUp()
        self.client = JSONAPITestClient()
        self.owner = BlueBottleUserFactory.create()
        self.random_user = BlueBottleUserFactory.create()
        self.supporter1 = BlueBottleUserFactory.create()
        self.supporter2 = BlueBottleUserFactory.create()
        self.activity = DateActivityFactory.create(
            review=False,
            owner=self.owner
        )
        self.slot = DateActivitySlotFactory.create(activity=self.activity)
        self.participant1 = DateParticipantFactory.create(
            user=self.supporter1,
            activity=self.activity
        )
        SlotParticipantFactory.create(slot=self.slot, participant=self.participant1)

        self.participant2 = DateParticipantFactory.create(
            user=self.supporter2,
            activity=self.activity
        )
        SlotParticipantFactory.create(slot=self.slot, participant=self.participant2)
        self.participant2.states.withdraw(save=True)

        p1_sl1 = SlotParticipant.objects.get(slot=self.slot, participant=self.participant1)
        p2_sl1 = SlotParticipant.objects.get(slot=self.slot, participant=self.participant2)
        self.url_part1_slot1 = reverse('slot-participant-detail', args=(p1_sl1.id,))
        self.url_part2_slot1 = reverse('slot-participant-detail', args=(p2_sl1.id,))

    def test_get_slot_participant(self):
        MemberPlatformSettings.objects.update(closed=True)
        response = self.client.get(self.url_part1_slot1, user=self.supporter1)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        response = self.client.get(self.url_part2_slot1, user=self.supporter1)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

        response = self.client.get(self.url_part1_slot1, user=self.owner)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

        response = self.client.get(self.url_part2_slot1, user=self.random_user)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


class SlotParticipantTransitionAPIViewTestCase(BluebottleTestCase):
    def setUp(self):
        super().setUp()
        self.client = JSONAPITestClient()
        self.activity = DateActivityFactory.create()
        self.slot = DateActivitySlotFactory.create(activity=self.activity)
        self.participant = DateParticipantFactory.create(activity=self.activity)
        self.slot_participant = SlotParticipantFactory.create(
            participant=self.participant, slot=self.slot
        )

        self.url = reverse('slot-participant-transition-list')
        self.data = {
            'data': {
                'type': 'contributors/time-based/slot-participant-transitions',
                'attributes': {},
                'relationships': {
                    'resource': {
                        'data': {
                            'type': 'contributors/time-based/slot-participants',
                            'id': self.slot_participant.pk
                        }
                    }
                }
            }
        }

    def test_withdraw_by_user(self):
        self.data['data']['attributes']['transition'] = 'withdraw'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.participant.user
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        data = json.loads(response.content)
        self.assertEqual(
            data['included'][2]['type'],
            'contributors/time-based/slot-participants'
        )
        self.assertEqual(data['included'][2]['meta']['status'], 'withdrawn')
        self.participant.refresh_from_db()
        contribution = self.participant.contributions.first()
        self.assertEqual(contribution.status, 'failed')

    def test_withdraw_by_user_with_preparation(self):
        self.activity = DateActivityFactory.create(slots=[])
        self.slot = DateActivitySlotFactory.create(activity=self.activity)
        self.activity.preparation = timedelta(hours=3)
        self.activity.save()
        self.participant = DateParticipantFactory.create(activity=self.activity)
        self.slot_participant = SlotParticipantFactory.create(
            participant=self.participant,
            slot=self.slot
        )
        self.data['data']['attributes']['transition'] = 'withdraw'
        self.data['data']['relationships']['resource']['data']['type'] = 'contributors/time-based/slot-participants'
        self.data['data']['relationships']['resource']['data']['id'] = self.slot_participant.pk

        contributions = self.participant.contributions.all()
        self.assertEqual(contributions[0].status, 'new')
        self.assertEqual(contributions[1].status, 'succeeded')

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.participant.user
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.participant.refresh_from_db()
        contributions = self.participant.contributions.all()
        self.assertEqual(contributions[0].status, 'failed')
        self.assertEqual(contributions[1].status, 'failed')

    def test_reapply_by_user(self):
        self.test_withdraw_by_user()

        self.data['data']['attributes']['transition'] = 'reapply'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.participant.user
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        data = json.loads(response.content)
        self.assertEqual(data['included'][2]['meta']['status'], 'registered')

    def test_reapply_by_user_with_prep_time(self):
        self.test_withdraw_by_user_with_preparation()

        self.data['data']['attributes']['transition'] = 'reapply'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.participant.user
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        data = json.loads(response.content)
        self.assertEqual(data['included'][2]['meta']['status'], 'registered')
        contributions = self.participant.contributions.all()
        self.assertEqual(contributions[0].status, 'new')
        self.assertEqual(contributions[1].status, 'succeeded')

    def test_withdraw_by_owner(self):
        self.data['data']['attributes']['transition'] = 'withdraw'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.activity.owner
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_reapply_by_owner(self):
        self.test_withdraw_by_user()

        self.data['data']['attributes']['transition'] = 'reapply'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.activity.owner
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_remove_by_owner(self):
        self.data['data']['attributes']['transition'] = 'remove'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.activity.owner
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        data = json.loads(response.content)
        self.assertEqual(data['included'][2]['meta']['status'], 'removed')

    def test_accept_by_owner(self):
        self.test_remove_by_owner()

        self.data['data']['attributes']['transition'] = 'accept'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.activity.owner
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        data = json.loads(response.content)
        self.assertEqual(data['included'][2]['meta']['status'], 'registered')

    def test_remove_by_user(self):
        self.data['data']['attributes']['transition'] = 'remove'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.participant.user
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_accept_by_user(self):
        self.test_remove_by_owner()
        self.data['data']['attributes']['transition'] = 'accept'

        response = self.client.post(
            self.url,
            json.dumps(self.data),
            user=self.participant.user
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)


class TimeContributionDetailAPIViewTestCase():
    def setUp(self):
        super().setUp()
        self.client = JSONAPITestClient()
        self.activity = self.factory.create()
        self.participant = self.participant_factory.create(
            activity=self.activity,
        )
        SlotParticipantFactory.create(
            slot=self.activity.slots.first(), participant=self.participant
        )
        self.contribution = self.participant.contributions.get()

        self.url = reverse(
            'time-contribution-detail',
            args=(self.contribution.pk,)
        )
        self.data = {
            'data': {
                'type': 'contributions/time-contributions',
                'id': self.contribution.pk,
                'attributes': {
                    'value': '5:00:00'
                }
            }
        }

    def test_get_owner(self):
        response = self.client.get(self.url, user=self.activity.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(
            response.json()['data']['meta']['permissions']['PUT']
        )

    def test_get_contributor(self):
        response = self.client.get(self.url, user=self.participant.user)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_get_other(self):
        response = self.client.get(self.url, user=BlueBottleUserFactory.create())

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_get_anonymous(self):
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_put_owner(self):
        response = self.client.put(self.url, json.dumps(self.data), user=self.activity.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.contribution.refresh_from_db()

        self.assertEqual(self.contribution.value, timedelta(hours=5))

    def test_put_contributor(self):
        response = self.client.put(self.url, json.dumps(self.data), user=self.participant.user)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_put_other(self):
        response = self.client.put(self.url, json.dumps(self.data), user=BlueBottleUserFactory.create())

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_put_anonymous(self):
        response = self.client.put(self.url, json.dumps(self.data))

        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)


class DateTimeContributionAPIViewTestCase(TimeContributionDetailAPIViewTestCase, BluebottleTestCase):
    factory = DateActivityFactory
    participant_factory = DateParticipantFactory


class SlotIcalTestCase(BluebottleTestCase):
    def setUp(self):
        super().setUp()
        self.user = BlueBottleUserFactory.create()
        self.client = JSONAPITestClient()
        self.initiative = InitiativeFactory.create(status='approved')
        self.activity = DateActivityFactory.create(
            title='Pollute Katwijk Beach',
            owner=self.user,
            initiative=self.initiative
        )
        self.slot = self.activity.slots.first()
        self.slot.is_online = True
        self.slot.online_meeting_url = 'http://example.com'
        self.slot.location = None
        self.slot.save()

        self.slot_url = reverse('date-slot-detail', args=(self.slot.pk,))
        self.activity.states.publish(save=True)
        self.client = JSONAPITestClient()
        response = self.client.get(self.slot_url, user=self.user)
        self.signed_url = response.json()['data']['attributes']['links']['ical']
        self.unsigned_url = reverse('slot-ical', args=(self.activity.pk,))

    def test_get(self):
        response = self.client.get(self.signed_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.get('content-type'), 'text/calendar')
        self.assertEqual(
            response.get('content-disposition'),
            'attachment; filename="{}.ics"'.format(self.activity.slug)
        )

        calendar = icalendar.Calendar.from_ical(response.content)

        for ical_event in calendar.walk('vevent'):
            self.assertAlmostEqual(
                ical_event['dtstart'].dt,
                self.slot.start,
                delta=timedelta(seconds=10)
            )
            self.assertAlmostEqual(
                ical_event['dtend'].dt,
                self.slot.start + self.slot.duration,
                delta=timedelta(seconds=10)
            )

            self.assertEqual(ical_event['dtstart'].dt.tzinfo, UTC)
            self.assertEqual(ical_event['dtend'].dt.tzinfo, UTC)

            self.assertEqual(str(ical_event['summary']), self.activity.title)
            self.assertEqual(
                str(ical_event['description']),
                '{}\nJoin: {}'.format(
                    self.activity.details,
                    self.slot.online_meeting_url
                )
            )
            self.assertEqual(ical_event['url'], self.activity.get_absolute_url())
            self.assertEqual(ical_event['organizer'], 'MAILTO:{}'.format(self.activity.owner.email))
            self.assertTrue('location' not in ical_event)

    def test_get_location(self):
        self.slot.location = GeolocationFactory.create()
        self.slot.save()

        response = self.client.get(self.signed_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        calendar = icalendar.Calendar.from_ical(response.content)

        for ical_event in calendar.walk('vevent'):
            self.assertEqual(ical_event['organizer'], 'MAILTO:{}'.format(self.activity.owner.email))
            self.assertEqual(
                ical_event['location'], self.slot.location.formatted_address
            )

    def test_get_location_hint(self):
        self.slot.location = GeolocationFactory.create()
        self.slot.location_hint = 'On the first floor'
        self.slot.save()

        response = self.client.get(self.signed_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        calendar = icalendar.Calendar.from_ical(response.content)

        for ical_event in calendar.walk('vevent'):
            self.assertEqual(ical_event['organizer'], 'MAILTO:{}'.format(self.activity.owner.email))
            self.assertEqual(
                ical_event['location'],
                f'{self.slot.location.formatted_address} ({self.slot.location_hint})'
            )

    def test_get_no_signature(self):
        response = self.client.get(self.unsigned_url)
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_get_wrong_signature(self):
        response = self.client.get('{}?signature=ewiorjewoijical_url'.format(self.unsigned_url))
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)


class DateIcalTestCase(BluebottleTestCase):
    def setUp(self):
        super().setUp()

        self.activity = DateActivityFactory.create(
            title='Pollute Katwijk Beach',
            slots=[]
        )
        self.slots = DateActivitySlotFactory.create_batch(
            3,
            activity=self.activity,
            is_online=True,
            online_meeting_url='http://example.com'
        )
        self.user = BlueBottleUserFactory.create()
        self.client = JSONAPITestClient()
        self.activity_url = reverse('date-detail', args=(self.activity.pk,))
        response = self.client.get(self.activity_url, user=self.user)
        self.signed_url = response.json()['data']['attributes']['links']['ical']
        self.unsigned_url = reverse('slot-ical', args=(self.activity.pk,))

    def test_get_applied_to_all(self):
        participant = DateParticipantFactory.create(
            activity=self.activity, user=self.user
        )
        for slot in self.activity.slots.all():
            SlotParticipantFactory.create(slot=slot, participant=participant)

        response = self.client.get(self.signed_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.assertEqual(response.get('content-type'), 'text/calendar')
        self.assertEqual(
            response.get('content-disposition'),
            'attachment; filename="{}.ics"'.format(self.activity.slug)
        )

        calendar = icalendar.Calendar.from_ical(response.content)
        self.assertEqual(len(calendar.walk('vevent')), 3)

        for index, ical_event in enumerate(calendar.walk('vevent')):
            slot = self.slots[index]
            self.assertAlmostEqual(
                ical_event['dtstart'].dt,
                slot.start,
                delta=timedelta(seconds=10)
            )
            self.assertAlmostEqual(
                ical_event['dtend'].dt,
                slot.start + slot.duration,
                delta=timedelta(seconds=10)
            )

            self.assertEqual(ical_event['dtstart'].dt.tzinfo, UTC)
            self.assertEqual(ical_event['dtend'].dt.tzinfo, UTC)

            self.assertEqual(str(ical_event['summary']), self.activity.title)
            self.assertEqual(
                str(ical_event['description']),
                '{}\nJoin: {}'.format(
                    self.activity.details,
                    slot.online_meeting_url
                )
            )
            self.assertEqual(ical_event['url'], self.activity.get_absolute_url())
            self.assertEqual(ical_event['organizer'], 'MAILTO:{}'.format(self.activity.owner.email))

    def test_get_applied_to_first(self):
        participant = DateParticipantFactory.create(activity=self.activity, user=self.user)
        SlotParticipantFactory.create(slot=self.slots[0], participant=participant)

        response = self.client.get(self.signed_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.assertEqual(response.get('content-type'), 'text/calendar')
        self.assertEqual(
            response.get('content-disposition'),
            'attachment; filename="{}.ics"'.format(self.activity.slug)
        )

        calendar = icalendar.Calendar.from_ical(response.content)
        self.assertEqual(len(calendar.walk('vevent')), 1)

        slot = self.slots[0]
        ical_event = list(calendar.walk('vevent'))[0]

        self.assertAlmostEqual(
            ical_event['dtstart'].dt,
            slot.start,
            delta=timedelta(seconds=10)
        )
        self.assertAlmostEqual(
            ical_event['dtend'].dt,
            slot.start + slot.duration,
            delta=timedelta(seconds=10)
        )

        self.assertEqual(ical_event['dtstart'].dt.tzinfo, UTC)
        self.assertEqual(ical_event['dtend'].dt.tzinfo, UTC)

        self.assertEqual(str(ical_event['summary']), self.activity.title)
        self.assertEqual(
            str(ical_event['description']),
            '{}\nJoin: {}'.format(
                self.activity.details,
                slot.online_meeting_url
            )
        )
        self.assertEqual(ical_event['url'], self.activity.get_absolute_url())
        self.assertEqual(ical_event['organizer'], 'MAILTO:{}'.format(self.activity.owner.email))

    def test_get_no_signature(self):
        response = self.client.get(self.unsigned_url)
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_get_wrong_signature(self):
        response = self.client.get('{}?signature=ewiorjewoijical_url'.format(self.unsigned_url))
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)


class SkillApiTestCase(BluebottleTestCase):

    def setUp(self):
        super().setUp()
        MemberPlatformSettings.objects.update(closed=True)
        self.url = reverse('skill-list')
        Skill.objects.all().delete()
        SkillFactory.create_batch(10)
        self.client = JSONAPITestClient()

    def test_get_skills_authenticated(self):
        user = BlueBottleUserFactory.create()
        response = self.client.get(self.url, user=user)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data['results']), 10)

    def test_get_skills_unauthenticated(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 401)

    def test_get_skills_old_url(self):
        old_url = reverse('assignment-skill-list')
        user = BlueBottleUserFactory.create()
        response = self.client.get(old_url, user=user)
        self.assertEqual(response.status_code, 200)


class RelatedSlotParticipantListViewTestCase(APITestCase):
    def setUp(self):
        self.client = JSONAPITestClient()

        self.activity = DateActivityFactory.create(slots=[])
        self.slots = DateActivitySlotFactory.create_batch(5, activity=self.activity)

        self.participant = DateParticipantFactory.create(activity=self.activity)

        self.slot_participants = [
            SlotParticipantFactory.create(participant=self.participant, slot=slot)
            for slot in self.slots[:3]
        ]

        self.slot_participants[0].states.remove(save=True)

        self.url = reverse('related-slot-participant-list', args=(self.participant.pk,))

    def test_get_user(self):
        self.perform_get(user=self.participant.user)

        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(3)
        self.assertIncluded('slot')

    def test_get_activity_owner(self):
        self.perform_get(user=self.activity.owner)
        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(3)

    def test_get_other_user(self):
        self.perform_get(user=BlueBottleUserFactory.create())

        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(2)

    def test_get_other_user_rejected_participant(self):
        self.participant.states.reject(save=True)
        self.perform_get(user=BlueBottleUserFactory.create())

        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(0)


class SlotRelatedParticipantListTestCase(APITestCase):
    def setUp(self):
        self.client = JSONAPITestClient()

        self.activity = DateActivityFactory.create(slots=[])
        self.slot = DateActivitySlotFactory.create(activity=self.activity)

        self.participants = DateParticipantFactory.create_batch(5, activity=self.activity)

        self.slot_participants = [
            SlotParticipantFactory.create(participant=participant, slot=self.slot)
            for participant in self.participants
        ]

        self.url = reverse('slot-participants', args=(self.slot.pk,))

    def test_get_user(self):
        self.perform_get(user=self.participants[0].user)

        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(5)

    def test_get_useri_only_firstname(self):
        MemberPlatformSettings.objects.update_or_create(display_member_names='first_name')
        self.perform_get(user=self.participants[0].user)

        self.assertStatus(status.HTTP_200_OK)

        for member in self.included_by_type(self.response, 'members'):
            self.assertIsNone(member['attributes']['last-name'])

    def test_get_activity_owner(self):
        self.perform_get(user=self.activity.owner)

        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(5)

    def test_get_activity_owner_only_first_name(self):
        MemberPlatformSettings.objects.update_or_create(display_member_names='first_name')
        self.perform_get(user=self.activity.owner)

        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(5)
        for member in self.included_by_type(self.response, 'members'):
            self.assertTrue(member['attributes']['last-name'])

    def test_get_staff_only_firstname(self):
        MemberPlatformSettings.objects.update_or_create(display_member_names='first_name')
        self.perform_get(user=BlueBottleUserFactory.create(is_staff=True))

        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(5)
        for member in self.included_by_type(self.response, 'members'):
            self.assertTrue(member['attributes']['last-name'])

    def test_get_activity_anonymous(self):
        self.perform_get()

        self.assertStatus(status.HTTP_200_OK)
        self.assertTotal(5)
