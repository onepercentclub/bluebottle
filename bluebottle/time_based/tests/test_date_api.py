import json
from datetime import date, timedelta
from io import BytesIO

import icalendar
from django.urls import reverse
from django.utils.timezone import now
from openpyxl import load_workbook
from rest_framework import status

from bluebottle.initiatives.models import InitiativePlatformSettings
from bluebottle.initiatives.tests.factories import InitiativeFactory
from bluebottle.members.models import MemberPlatformSettings
from bluebottle.test.factory_models.accounts import BlueBottleUserFactory
from bluebottle.test.utils import APITestCase, JSONAPITestClient
from bluebottle.time_based.serializers import (
    DateActivitySerializer,
    DateParticipantSerializer,
    DateParticipantTransitionSerializer,
    DateRegistrationSerializer,
    DateRegistrationTransitionSerializer,
    DateTransitionSerializer,
    DateActivitySlotSerializer,
)
from bluebottle.time_based.tests.base import (
    TimeBasedActivityAPIExportTestCase,
    TimeBasedActivityDetailAPITestCase,
    TimeBasedActivityListAPITestCase,
    TimeBasedActivityTransitionListAPITestCase,
    TimeBasedParticipantDetailAPITestCase,
    TimeBasedParticipantRelatedListAPITestCase,
    TimeBasedParticipantTransitionListAPITestCase,
    TimeBasedRegistrationDetailAPITestCase,
    TimeBasedRegistrationListAPITestCase,
    TimeBasedRegistrationRelatedAPIListTestCase,
    TimeBasedRegistrationTransitionListAPITestCase,
)
from bluebottle.time_based.tests.factories import (
    DateActivityFactory,
    DateParticipantFactory,
    DateRegistrationFactory,
    DateActivitySlotFactory,
)


class DateActivityListAPITestCase(TimeBasedActivityListAPITestCase, APITestCase):
    url_name = 'date-list'
    serializer = DateActivitySerializer
    factory = DateActivityFactory

    fields = ['initiative', 'title', 'description', 'review', 'theme']
    relationships = ['initiative', 'owner', 'theme']
    attributes = ['title', 'description', 'review']
    included = ['initiative', 'owner', 'theme']

    defaults = {
        'title': 'Test title',
        'description': json.dumps({'html': 'Test description', 'delta': ''}),
        'review': False,
    }

    def setUp(self):
        super().setUp()

    def test_create_complete(self, user=None, data=None):
        user = self.defaults['initiative'].owner
        self.perform_create(user=user)
        self.assertStatus(status.HTTP_201_CREATED)

        for relationship in self.relationships:
            self.assertRelationship(relationship)

        for included in self.included:
            self.assertIncluded(included)

        for attribute in self.attributes:
            self.assertAttribute(attribute)

        self.assertPermission('PUT', True)
        self.assertPermission('GET', True)
        self.assertPermission('PATCH', True)

        self.assertTransition('delete')

        activity_id = self.response.json()['data']['id']
        slot_url = reverse('date-slot-list')
        response = self.client.post(
            slot_url,
            {
                'data': {
                    'type': 'activities/time-based/date-slots',
                    'attributes': {
                        'start': '2026-01-01 10:00:00',
                        'duration': '01:00',
                        'is-online': True,
                        'title': ''
                    },
                    'relationships': {
                        'activity': {
                            'data': {
                                'id': activity_id,
                                'type': 'activities/time-based/dates'
                            }
                        }
                    }
                }
            },
            HTTP_AUTHORIZATION="JWT {0}".format(user.get_jwt_token())
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        activity = [
            resource for resource in response.json()['included']
            if resource['type'] == 'activities/time-based/dates'
        ][0]

        self.assertTrue(
            'publish' in [transition['name'] for transition in activity['meta']['transitions']]
        )


class DateActivityDetailAPITestCase(TimeBasedActivityDetailAPITestCase, APITestCase):
    url_name = 'date-detail'
    serializer = DateActivitySerializer
    factory = DateActivityFactory

    fields = ['initiative', 'title', 'description', 'review']
    attributes = ['title', 'description', 'review']

    defaults = {
        'title': 'Test title',
        'description': json.dumps({'html': 'Test description', 'delta': ''}),
        'review': False,
    }

    def test_put_start_after_end(self):
        pass

    def test_get_multiple_slots(self):
        self.model.slots.get().delete()

        DateActivitySlotFactory.create_batch(3, activity=self.model, status='open', capacity=10)
        DateActivitySlotFactory.create_batch(3, activity=self.model, status='cancelled', capacity=10)

        self.perform_get(user=self.model.owner)

        self.assertEqual(self.response.json()['data']['attributes']['date-info']['capacity'], 30)


class DateActivityTransitionListAPITestCase(TimeBasedActivityTransitionListAPITestCase, APITestCase):
    url_name = 'date-transition-list'
    serializer = DateTransitionSerializer
    activity_factory = DateActivityFactory
    fields = ['resource', 'transition']
    defaults = {
        'title': 'Test title',
        'description': json.dumps({'html': 'Test description', 'delta': ''}),
        'review': False,
    }


class DateRegistrationListAPITestCase(TimeBasedRegistrationListAPITestCase, APITestCase):
    url_name = 'date-registration-list'
    serializer = DateRegistrationSerializer
    factory = DateRegistrationFactory
    activity_factory = DateActivityFactory

    activity_defaults = {}


class DateRegistrationRelatedListAPITestCase(TimeBasedRegistrationRelatedAPIListTestCase, APITestCase):
    url_name = 'related-date-registrations'
    serializer = DateRegistrationSerializer
    factory = DateRegistrationFactory
    activity_factory = DateActivityFactory

    activity_defaults = {}


class DateRegistrationDetailAPITestCase(TimeBasedRegistrationDetailAPITestCase, APITestCase):
    url_name = 'date-registration-detail'
    serializer = DateRegistrationSerializer
    factory = DateRegistrationFactory
    activity_factory = DateActivityFactory

    activity_defaults = {}


class DateRegistrationTransitionListAPITestCase(TimeBasedRegistrationTransitionListAPITestCase, APITestCase):
    url_name = 'date-registration-transitions'
    serializer = DateRegistrationTransitionSerializer

    factory = DateRegistrationFactory
    activity_factory = DateActivityFactory
    activity_defaults = {}


class DateParticipantRelatedListAPITestCase(TimeBasedParticipantRelatedListAPITestCase, APITestCase):
    url_name = 'date-participants'
    serializer = DateParticipantSerializer
    factory = DateParticipantFactory

    activity_factory = DateActivityFactory

    activity_defaults = {}


class DateParticipantDetailAPITestCase(TimeBasedParticipantDetailAPITestCase, APITestCase):
    url_name = 'date-participant-detail'
    serializer = DateParticipantSerializer
    factory = DateParticipantFactory
    activity_factory = DateActivityFactory

    activity_defaults = {}


class DateParticipantTransitionListAPITestCase(TimeBasedParticipantTransitionListAPITestCase, APITestCase):
    url_name = 'date-participant-transitions'
    serializer = DateParticipantTransitionSerializer

    factory = DateParticipantFactory
    activity_factory = DateActivityFactory

    activity_defaults = {}

    transition = 'remove'
    target_status = 'removed'


class DateActivityExportTestCase(TimeBasedActivityAPIExportTestCase, APITestCase):
    factory = DateActivityFactory
    participant_factory = DateParticipantFactory
    url_name = 'date-detail'

    activity_defaults = {}

    def test_get(self):
        self.perform_get(user=self.activity.owner)

        self.assertStatus(status.HTTP_200_OK)

        workbook = load_workbook(filename=BytesIO(self.response.content))
        self.assertEqual(len(workbook.worksheets), 5)

        sheet = workbook.get_active_sheet()

        self.assertEqual(
            tuple(sheet.values)[0],
            ('Email', 'Name', 'Registration Date', 'Status', 'Registration answer',)
        )


class DateSlotDetailAPITestCase(APITestCase):
    url_name = 'date-slot-detail'
    serializer = DateActivitySlotSerializer
    factory = DateActivitySlotFactory

    fields = []
    attributes = [
        'start', 'duration', 'is-online'
    ]
    included = ['activity']

    defaults = {}

    def setUp(self):
        super().setUp()
        self.manager = BlueBottleUserFactory.create()
        self.admin = BlueBottleUserFactory.create(is_staff=True)
        self.user = BlueBottleUserFactory.create()
        self.participant = BlueBottleUserFactory.create()
        self.activity = DateActivityFactory.create(
            initiative=InitiativeFactory.create(status='approved'),
            status='open',
            review=False,
            owner=self.manager,
        )

        self.model = self.factory.create(activity=self.activity)
        self.url = reverse(self.url_name, args=(self.model.pk,))

    def test_get(self):
        self.perform_get(user=self.manager)
        self.assertStatus(status.HTTP_200_OK)

        for attribute in self.attributes:
            self.assertAttribute(attribute)

        for relationship in self.included:
            self.assertIncluded(relationship)

    def test_get_anonymous(self):
        self.perform_get()
        self.assertStatus(status.HTTP_200_OK)

        for attribute in self.attributes:
            self.assertAttribute(attribute)

        for relationship in self.included:
            self.assertIncluded(relationship)

    def test_update_activity_manager(self):
        start = (date.today() + timedelta(days=10)).strftime('%Y-%m-%d %H:00:00')
        self.perform_update({'start': start, 'duration': '4:0:0'}, user=self.manager)
        self.assertStatus(status.HTTP_200_OK)

    def test_update_admin(self):
        start = (date.today() + timedelta(days=10)).strftime('%Y-%m-%d %H:00:00')
        self.perform_update({'start': start, 'duration': '4:0:0'}, user=self.admin)
        self.assertStatus(status.HTTP_200_OK)

    def test_set_update_participant(self):
        start = (date.today() + timedelta(days=10)).strftime('%Y-%m-%d %H:00:00')
        self.perform_update({'start': start, 'duration': '4:0:0'}, user=self.participant)
        self.assertStatus(status.HTTP_403_FORBIDDEN)

    def test_set_date_user(self):
        start = (date.today() + timedelta(days=10)).strftime('%Y-%m-%d %H:00:00')
        self.perform_update({'start': start, 'duration': '4:0:0'}, user=self.user)
        self.assertStatus(status.HTTP_403_FORBIDDEN)

    def test_ical_download(self):
        start = (date.today() + timedelta(days=10)).strftime("%Y-%m-%d %H:00:00")
        self.perform_update(
            {"start": start, "duration": "4:0:0", "is_online": True}, user=self.admin
        )

        ical_response = self.client.get(
            self.response.json()["data"]["attributes"]["links"]["ical"],
        )

        calendar = icalendar.Calendar.from_ical(ical_response.content)

        for ical_event in calendar.walk("vevent"):
            self.assertAlmostEqual(
                ical_event["dtstart"].dt, self.model.start, delta=timedelta(seconds=10)
            )
            self.assertAlmostEqual(
                ical_event["dtend"].dt,
                self.model.start + self.model.duration,
                delta=timedelta(seconds=10),
            )

            self.assertEqual(str(ical_event["summary"]), self.activity.title)
            self.assertEqual(ical_event["url"], self.activity.get_absolute_url())
            self.assertEqual(
                ical_event["organizer"], "MAILTO:{}".format(self.activity.owner.email)
            )

    def test_export_download_anonymous(self):
        settings = InitiativePlatformSettings.objects.get()
        settings.enable_participant_exports = True
        settings.save()

        self.perform_get()
        self.assertIsNone(self.response.json()['data']['attributes']['participants-export-url'])

    def test_export_download_disabled(self):
        self.perform_get(user=self.model.owner)
        self.assertIsNone(self.response.json()['data']['attributes']['participants-export-url'])

    def test_export_download_owner(self):
        settings = InitiativePlatformSettings.objects.get()
        settings.enable_participant_exports = True
        settings.save()

        self.perform_get(user=self.model.owner)

        self.assertTrue(
            self.response.json()['data']['attributes']['participants-export-url']['url'].startswith(
                reverse('slot-participant-export', args=(self.model.pk,))
            ),
        )

        export_response = self.client.get(
            self.response.json()['data']['attributes']['participants-export-url']['url']
        )
        workbook = load_workbook(filename=BytesIO(export_response.content))
        self.assertEqual(len(workbook.worksheets), 1)

        sheet = workbook.get_active_sheet()

        self.assertEqual(
            tuple(sheet.values)[0],
            ('Email', 'Name', 'Registration Date', 'Status',)
        )


class DateSlotListAPITestCase(APITestCase):
    url_name = 'date-slot-list'
    serializer = DateActivitySlotSerializer
    factory = DateActivitySlotFactory

    fields = ['activity', 'duration', 'is_online']
    attributes = [
        'start', 'duration', 'is_online'
    ]
    included = ['activity']

    def setUp(self):
        super().setUp()
        self.manager = BlueBottleUserFactory.create()
        self.admin = BlueBottleUserFactory.create(is_staff=True)
        self.user = BlueBottleUserFactory.create()
        self.participant = BlueBottleUserFactory.create()
        self.activity = DateActivityFactory.create(
            initiative=InitiativeFactory.create(status='approved'),
            status='open',
            review=False,
            owner=self.manager,
        )

        self.url = reverse(self.url_name)

        self.defaults = {
            'is_online': False,
            'activity': self.activity
        }

    def test_get(self):
        self.perform_get(user=self.manager)
        self.assertStatus(status.HTTP_405_METHOD_NOT_ALLOWED)

    def test_create(self):
        self.perform_create(user=self.activity.owner)
        self.assertStatus(status.HTTP_201_CREATED)

        for included in self.included:
            self.assertIncluded(included)

        for attribute in self.attributes:
            self.assertAttribute(attribute)

        self.assertPermission('PUT', True)
        self.assertPermission('GET', True)
        self.assertPermission('PATCH', True)

    def test_create_manager(self):
        self.perform_create(user=self.manager)
        self.assertStatus(status.HTTP_201_CREATED)

    def test_create_admin(self):
        self.perform_create(user=self.manager)
        self.assertStatus(status.HTTP_201_CREATED)

    def test_create_other_user(self):
        self.perform_create(user=BlueBottleUserFactory.create())
        self.assertStatus(status.HTTP_403_FORBIDDEN)

    def test_create_anonymous(self):
        self.perform_create()
        self.assertStatus(status.HTTP_401_UNAUTHORIZED)


class DateSlotRelatedListAPITestCase(APITestCase):
    url_name = 'related-date-slots'
    serializer = DateActivitySlotSerializer
    factory = DateActivitySlotFactory

    attributes = [
        'start', 'duration', 'is_online'
    ]
    included = ['activity']

    def setUp(self):
        super().setUp()
        self.manager = BlueBottleUserFactory.create()
        self.admin = BlueBottleUserFactory.create(is_staff=True)
        self.user = BlueBottleUserFactory.create()
        self.participant = BlueBottleUserFactory.create()
        self.activity = DateActivityFactory.create(
            initiative=InitiativeFactory.create(status='approved'),
            status='open',
            review=False,
            owner=self.manager,
            slots=[]
        )

        self.factory.create_batch(3, activity=self.activity, start=now() + timedelta(days=5))
        self.factory.create_batch(2, activity=self.activity, start=now() - timedelta(days=5))
        self.url = reverse(self.url_name, args=(self.activity.pk,))

    def test_get_manager_future(self):
        self.perform_get(user=self.manager)
        self.assertStatus(status.HTTP_200_OK)

        self.assertTotal(5)

        for included in self.included:
            self.assertIncluded(included)

        for attribute in self.attributes:
            self.assertAttribute(attribute)

    def test_get_future(self):
        self.perform_get(query={'start': date.today(), 'ordering': 'start'})
        self.assertStatus(status.HTTP_200_OK)

        self.assertTotal(3)

        for included in self.included:
            self.assertIncluded(included)

        for attribute in self.attributes:
            self.assertAttribute(attribute)

    def test_get_passed(self):
        self.perform_get(query={'start': date.today(), 'ordering': '-start'})
        self.assertStatus(status.HTTP_200_OK)

        self.assertTotal(2)

        for included in self.included:
            self.assertIncluded(included)

        for attribute in self.attributes:
            self.assertAttribute(attribute)

    def test_get_other_user(self):
        self.perform_get(user=BlueBottleUserFactory.create())
        self.assertStatus(status.HTTP_200_OK)

    def test_get_anonymous(self):
        self.perform_get()
        self.assertStatus(status.HTTP_200_OK)


class DateSlotRelatedParticipantsListAPITestCase(APITestCase):
    url_name = 'date-slot-related-participants'
    serializer = DateParticipantSerializer
    factory = DateParticipantFactory

    attributes = []
    included = ['slot', 'activity']

    def setUp(self):
        super().setUp()
        self.manager = BlueBottleUserFactory.create()
        self.admin = BlueBottleUserFactory.create(is_staff=True)
        self.user = BlueBottleUserFactory.create()
        self.participant = BlueBottleUserFactory.create()
        self.activity = DateActivityFactory.create(
            initiative=InitiativeFactory.create(status='approved'),
            status='open',
            review=False,
            owner=self.manager,
            slots=[]
        )
        self.slot = DateActivitySlotFactory.create(activity=self.activity)
        self.factory.create_batch(2, slot=self.slot, status='accepted')
        self.factory.create_batch(2, slot=self.slot, status='succeeded')
        self.factory.create_batch(2, slot=self.slot, status='withdrawn')

        self.url = reverse(self.url_name, args=(self.slot.pk,))

    def test_get_manager(self):
        self.perform_get(user=self.manager)
        self.assertStatus(status.HTTP_200_OK)

        self.assertTotal(6)

        for included in self.included:
            self.assertIncluded(included)

        for attribute in self.attributes:
            self.assertAttribute(attribute)

    def test_get_admin(self):
        self.perform_get(user=self.admin)
        self.assertStatus(status.HTTP_200_OK)

        self.assertTotal(6)

    def test_other_user(self):
        self.perform_get(user=BlueBottleUserFactory.create())
        self.assertStatus(status.HTTP_200_OK)

        self.assertTotal(4)

        for resource in self.response.json()['data']:
            self.assertTrue(
                resource['meta']['current-status']['value'] in ['accepted', 'succeeded']
            )

    def test_failed_participant(self):
        participant = self.slot.participants.filter(status='withdrawn').first()

        self.perform_get(user=participant.user)
        self.assertStatus(status.HTTP_200_OK)

        self.assertTotal(5)

        for resource in self.response.json()['data']:
            self.assertTrue(
                resource['meta']['current-status']['value'] in ['accepted', 'succeeded', 'withdrawn']
            )

    def test_anonymous(self):
        self.perform_get()
        self.assertStatus(status.HTTP_200_OK)

        self.assertTotal(4)

        for resource in self.response.json()['data']:
            self.assertTrue(
                resource['meta']['current-status']['value'] in ['accepted', 'succeeded']
            )

    def test_get_user_only_firstname(self):
        MemberPlatformSettings.objects.update_or_create(display_member_names='first_name')
        self.perform_get(user=BlueBottleUserFactory.create())

        self.assertStatus(status.HTTP_200_OK)
        for member in self.included_by_type(self.response, 'members'):
            self.assertIsNone(member['attributes']['last-name'])

    def test_get_activity_owner_only_first_name(self):
        MemberPlatformSettings.objects.update_or_create(display_member_names='first_name')
        self.perform_get(user=self.activity.owner)

        self.assertStatus(status.HTTP_200_OK)
        for member in self.included_by_type(self.response, 'members'):
            self.assertTrue(member['attributes']['last-name'])

    def test_get_staff_only_firstname(self):
        MemberPlatformSettings.objects.update_or_create(display_member_names='first_name')
        self.perform_get(user=BlueBottleUserFactory.create(is_staff=True))

        self.assertStatus(status.HTTP_200_OK)
        for member in self.included_by_type(self.response, 'members'):
            self.assertTrue(member['attributes']['last-name'])


class DateActivitySpotsLeftAPITestCase(APITestCase):
    """Test cases to verify spots_left calculation in date_info"""

    def setUp(self):
        super().setUp()
        self.client = JSONAPITestClient()
        self.owner = BlueBottleUserFactory.create()
        self.initiative = InitiativeFactory.create(status='approved', owner=self.owner)

        settings = InitiativePlatformSettings.objects.get()
        settings.activity_types.append('dateactivity')
        settings.save()

    def test_spots_left_with_upcoming_slot_with_capacity(self):
        """Test spots_left calculation with one upcoming slot with capacity set"""
        activity = DateActivityFactory.create(
            owner=self.owner,
            initiative=self.initiative,
            status='open'
        )

        # Delete default slot
        activity.slots.all().delete()

        # Create an upcoming slot with capacity
        slot = DateActivitySlotFactory.create(
            activity=activity,
            capacity=10,
            start=now() + timedelta(days=7),
            status='open'
        )

        # Add 3 accepted participants
        for _ in range(3):
            registration = DateRegistrationFactory.create(activity=activity, status='accepted')
            DateParticipantFactory.create(
                activity=activity,
                slot=slot,
                registration=registration,
                status='accepted'
            )

        url = reverse('date-detail', args=(activity.pk,))
        response = self.client.get(url, user=self.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        date_info = response.json()['data']['attributes']['date-info']

        self.assertEqual(date_info['capacity'], 10)
        self.assertEqual(date_info['spots_left'], 7)  # 10 - 3 = 7

    def test_spots_left_with_upcoming_slot_without_capacity(self):
        """Test spots_left is None when upcoming slot has no capacity"""
        activity = DateActivityFactory.create(
            owner=self.owner,
            initiative=self.initiative,
            status='open'
        )

        # Delete default slot
        activity.slots.all().delete()

        # Create an upcoming slot without capacity
        slot = DateActivitySlotFactory.create(
            activity=activity,
            capacity=None,
            start=now() + timedelta(days=7),
            status='open'
        )

        # Add participants
        for _ in range(2):
            registration = DateRegistrationFactory.create(activity=activity, status='accepted')
            DateParticipantFactory.create(
                activity=activity,
                slot=slot,
                registration=registration,
                status='accepted'
            )

        url = reverse('date-detail', args=(activity.pk,))
        response = self.client.get(url, user=self.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        date_info = response.json()['data']['attributes']['date-info']

        self.assertIsNone(date_info['capacity'])
        self.assertIsNone(date_info['spots_left'])

    def test_spots_left_with_past_slot_with_capacity(self):
        """Test spots_left calculation ignores past slots"""
        activity = DateActivityFactory.create(
            owner=self.owner,
            initiative=self.initiative,
            status='open'
        )

        # Delete default slot
        activity.slots.all().delete()

        # Create a past slot with capacity and participants
        past_slot = DateActivitySlotFactory.create(
            activity=activity,
            capacity=10,
            start=now() - timedelta(days=7),
            status='succeeded'
        )

        # Add 5 participants to past slot
        for _ in range(5):
            registration = DateRegistrationFactory.create(activity=activity, status='accepted')
            DateParticipantFactory.create(
                activity=activity,
                slot=past_slot,
                registration=registration,
                status='succeeded'
            )

        url = reverse('date-detail', args=(activity.pk,))
        response = self.client.get(url, user=self.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        date_info = response.json()['data']['attributes']['date-info']

        # Past slots should not affect spots_left for upcoming slots
        # Since there are no upcoming slots, capacity and spots_left should be None
        self.assertIsNone(date_info['capacity'])
        self.assertIsNone(date_info['spots_left'])

    def test_spots_left_with_multiple_upcoming_slots_all_with_capacity(self):
        """Test spots_left calculation with multiple upcoming slots all with capacity"""
        activity = DateActivityFactory.create(
            owner=self.owner,
            initiative=self.initiative,
            status='open'
        )

        # Delete default slot
        activity.slots.all().delete()

        # Create multiple upcoming slots with capacity
        slot1 = DateActivitySlotFactory.create(
            activity=activity,
            capacity=10,
            start=now() + timedelta(days=7),
            status='open'
        )
        slot2 = DateActivitySlotFactory.create(
            activity=activity,
            capacity=15,
            start=now() + timedelta(days=14),
            status='open'
        )

        # Add participants to slot1
        for _ in range(3):
            registration = DateRegistrationFactory.create(activity=activity, status='accepted')
            DateParticipantFactory.create(
                activity=activity,
                slot=slot1,
                registration=registration,
                status='accepted'
            )

        # Add participants to slot2
        for _ in range(5):
            registration = DateRegistrationFactory.create(activity=activity, status='accepted')
            DateParticipantFactory.create(
                activity=activity,
                slot=slot2,
                registration=registration,
                status='accepted'
            )

        url = reverse('date-detail', args=(activity.pk,))
        response = self.client.get(url, user=self.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        date_info = response.json()['data']['attributes']['date-info']

        self.assertEqual(date_info['capacity'], 25)  # 10 + 15 = 25
        self.assertEqual(date_info['spots_left'], 17)  # 25 - (3 + 5) = 17

    def test_spots_left_with_multiple_slots_mixed_capacity(self):
        """Test spots_left is None when at least one upcoming slot has no capacity"""
        activity = DateActivityFactory.create(
            owner=self.owner,
            initiative=self.initiative,
            status='open'
        )

        # Delete default slot
        activity.slots.all().delete()

        # Create slots with mixed capacity
        slot1 = DateActivitySlotFactory.create(
            activity=activity,
            capacity=10,
            start=now() + timedelta(days=7),
            status='open'
        )
        DateActivitySlotFactory.create(
            activity=activity,
            capacity=None,  # No capacity
            start=now() + timedelta(days=14),
            status='open'
        )

        # Add participants
        for _ in range(3):
            registration = DateRegistrationFactory.create(activity=activity, status='accepted')
            DateParticipantFactory.create(
                activity=activity,
                slot=slot1,
                registration=registration,
                status='accepted'
            )

        url = reverse('date-detail', args=(activity.pk,))
        response = self.client.get(url, user=self.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        date_info = response.json()['data']['attributes']['date-info']

        # If any slot has no capacity, capacity and spots_left should be None
        self.assertIsNone(date_info['capacity'])
        self.assertIsNone(date_info['spots_left'])

    def test_spots_left_with_pending_participants(self):
        """Test that pending participants (status='new') are not counted in spots_left"""
        activity = DateActivityFactory.create(
            owner=self.owner,
            initiative=self.initiative,
            status='open'
        )

        # Delete default slot
        activity.slots.all().delete()

        # Create an upcoming slot with capacity
        slot = DateActivitySlotFactory.create(
            activity=activity,
            capacity=10,
            start=now() + timedelta(days=7),
            status='open'
        )

        # Add 2 accepted participants
        for _ in range(2):
            registration = DateRegistrationFactory.create(activity=activity, status='accepted')
            DateParticipantFactory.create(
                activity=activity,
                slot=slot,
                registration=registration,
                status='accepted'
            )

        # Add 3 pending participants (should not count)
        for _ in range(3):
            registration = DateRegistrationFactory.create(activity=activity, status='new')
            DateParticipantFactory.create(
                activity=activity,
                slot=slot,
                registration=registration,
                status='new'
            )

        url = reverse('date-detail', args=(activity.pk,))
        response = self.client.get(url, user=self.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        date_info = response.json()['data']['attributes']['date-info']

        self.assertEqual(date_info['capacity'], 10)
        # Only accepted participants should be counted
        self.assertEqual(date_info['spots_left'], 8)  # 10 - 2 = 8

    def test_spots_left_with_rejected_and_withdrawn_participants(self):
        """Test that rejected and withdrawn participants are not counted"""
        activity = DateActivityFactory.create(
            owner=self.owner,
            initiative=self.initiative,
            status='open'
        )

        # Delete default slot
        activity.slots.all().delete()

        # Create an upcoming slot with capacity
        slot = DateActivitySlotFactory.create(
            activity=activity,
            capacity=10,
            start=now() + timedelta(days=7),
            status='open'
        )

        # Add 3 accepted participants
        for _ in range(3):
            registration = DateRegistrationFactory.create(activity=activity, status='accepted')
            DateParticipantFactory.create(
                activity=activity,
                slot=slot,
                registration=registration,
                status='accepted'
            )

        # Add rejected participants (should not count)
        for _ in range(2):
            registration = DateRegistrationFactory.create(activity=activity, status='rejected')
            DateParticipantFactory.create(
                activity=activity,
                slot=slot,
                registration=registration,
                status='rejected'
            )

        # Add withdrawn participant (should not count)
        registration = DateRegistrationFactory.create(activity=activity, status='withdrawn')
        DateParticipantFactory.create(
            activity=activity,
            slot=slot,
            registration=registration,
            status='withdrawn'
        )

        url = reverse('date-detail', args=(activity.pk,))
        response = self.client.get(url, user=self.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        date_info = response.json()['data']['attributes']['date-info']

        self.assertEqual(date_info['capacity'], 10)
        # Only accepted participants should be counted
        self.assertEqual(date_info['spots_left'], 7)  # 10 - 3 = 7

    def test_spots_left_with_succeeded_participants(self):
        """Test that succeeded participants are counted in spots_left"""
        activity = DateActivityFactory.create(
            owner=self.owner,
            initiative=self.initiative,
            status='open'
        )

        # Delete default slot
        activity.slots.all().delete()

        # Create an upcoming slot with capacity
        slot = DateActivitySlotFactory.create(
            activity=activity,
            capacity=10,
            start=now() + timedelta(days=7),
            status='open'
        )

        # Add 2 accepted participants
        for _ in range(2):
            registration = DateRegistrationFactory.create(activity=activity, status='accepted')
            DateParticipantFactory.create(
                activity=activity,
                slot=slot,
                registration=registration,
                status='accepted'
            )

        # Add 3 succeeded participants
        for _ in range(3):
            registration = DateRegistrationFactory.create(activity=activity, status='accepted')
            DateParticipantFactory.create(
                activity=activity,
                slot=slot,
                registration=registration,
                status='succeeded'
            )

        url = reverse('date-detail', args=(activity.pk,))
        response = self.client.get(url, user=self.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        date_info = response.json()['data']['attributes']['date-info']

        self.assertEqual(date_info['capacity'], 10)
        # Both accepted and succeeded participants should be counted
        self.assertEqual(date_info['spots_left'], 5)  # 10 - (2 + 3) = 5

    def test_spots_left_complex_scenario(self):
        """Test complex scenario with past and upcoming slots, mixed capacities and participant statuses"""
        activity = DateActivityFactory.create(
            owner=self.owner,
            initiative=self.initiative,
            status='open'
        )

        # Delete default slot
        activity.slots.all().delete()

        # Create a past slot with capacity (should be ignored)
        past_slot = DateActivitySlotFactory.create(
            activity=activity,
            capacity=20,
            start=now() - timedelta(days=7),
            status='succeeded'
        )

        # Add participants to past slot (should be ignored)
        for _ in range(10):
            registration = DateRegistrationFactory.create(activity=activity, status='accepted')
            DateParticipantFactory.create(
                activity=activity,
                slot=past_slot,
                registration=registration,
                status='succeeded'
            )

        # Create upcoming slot 1 with capacity
        upcoming_slot1 = DateActivitySlotFactory.create(
            activity=activity,
            capacity=15,
            start=now() + timedelta(days=7),
            status='open'
        )

        # Create upcoming slot 2 with capacity
        upcoming_slot2 = DateActivitySlotFactory.create(
            activity=activity,
            capacity=10,
            start=now() + timedelta(days=14),
            status='open'
        )

        # Add 5 accepted participants to upcoming_slot1
        for _ in range(5):
            registration = DateRegistrationFactory.create(activity=activity, status='accepted')
            DateParticipantFactory.create(
                activity=activity,
                slot=upcoming_slot1,
                registration=registration,
                status='accepted'
            )

        # Add 3 succeeded participants to upcoming_slot2
        for _ in range(3):
            registration = DateRegistrationFactory.create(activity=activity, status='accepted')
            DateParticipantFactory.create(
                activity=activity,
                slot=upcoming_slot2,
                registration=registration,
                status='succeeded'
            )

        # Add 2 pending participants to upcoming_slot1 (should not count)
        for _ in range(2):
            registration = DateRegistrationFactory.create(activity=activity, status='new')
            DateParticipantFactory.create(
                activity=activity,
                slot=upcoming_slot1,
                registration=registration,
                status='new'
            )

        url = reverse('date-detail', args=(activity.pk,))
        response = self.client.get(url, user=self.owner)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        date_info = response.json()['data']['attributes']['date-info']

        # Total capacity of upcoming slots only
        self.assertEqual(date_info['capacity'], 25)  # 15 + 10 = 25
        # Only accepted and succeeded participants in upcoming slots
        self.assertEqual(date_info['spots_left'], 17)  # 25 - (5 + 3) = 17
